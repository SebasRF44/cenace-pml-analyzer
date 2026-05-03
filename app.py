"""
CENACE PML Analyzer — Streamlit Cloud Edition  v9
======================================================
Recurrent Energy / Canadian Solar — SRF · Sebastian Roldan

v9 changes vs v8:
- Auto-detect Sistema (SIN/BCA/BCS) desde catálogo, queda solo MDA/MTR en sidebar
- Default fechas: hoy −2 semanas → hoy −1 día (delay CENACE)
- Warning si fecha fin > hoy −7 días (delay típico)
- "Moneda del análisis" → "FX Selector"
- Título "⚡🔋 Node Analyzer" más grande
- Multi-año en Excel: 1 hoja por nodo si ≤20 nodos (Opción C)
- Spread por CCR: gráfica continua con líneas por zona (≥2 CCRs)
- Excel Custom con toggle de gráficos seleccionables
- Click en filtro limpia consulta anterior
- Botón "Limpiar todo" para reset completo
- Badge "X nodos del filtro"
- Bug leyenda USD/MWh arreglado
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
import json
import urllib3
import time
import io
import os
import re
import gc
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, date
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from difflib import SequenceMatcher

import plotly.express as px
import plotly.graph_objects as go

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ═══════════════════════════════════════════════════════════════════════
# CONFIG STREAMLIT
# ═══════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="CENACE PML Analyzer · Recurrent Energy",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Workers fijo (no más slider)
MAX_WORKERS = 8

# ═══════════════════════════════════════════════════════════════════════
# COLORES
# ═══════════════════════════════════════════════════════════════════════
RE_NAVY     = "#0e346b"
RE_RED      = "#a0090c"
RE_BLUE     = "#2777bd"
RE_ALT      = "#EBF3FB"
RE_INFO     = "#D9E2F3"
TEXT_DARK   = "#1a1a1a"
TEXT_TITLE  = "#0a2347"
GRID_LIGHT  = "#D8D8D8"
AXIS_LINE   = "#444444"

PALETTE = [
    "#0e346b", "#a0090c", "#d4a017", "#1a8a3a", "#7e57c2",
    "#ff5722", "#00897b", "#c2185b", "#3949ab", "#5d4037",
    "#1976d2", "#558b2f", "#f57c00", "#8e24aa", "#455a64",
]


# ═══════════════════════════════════════════════════════════════════════
# MONEDA — helpers para el selector MXN / USD
# ═══════════════════════════════════════════════════════════════════════
def get_moneda():
    """Retorna 'MXN' o 'USD' según session_state."""
    return st.session_state.get("moneda", "MXN")


def simbolo_moneda():
    """Retorna '$' (MXN) o 'USD$' para diferenciarlos visualmente."""
    return "USD$" if get_moneda() == "USD" else "$"


def fmt_moneda():
    """Retorna formato Streamlit dataframe column compatible: '$%.2f' o 'USD$%.2f'."""
    if get_moneda() == "USD":
        return "USD$%.2f"
    return "$%.2f"


def label_moneda():
    """Retorna label completo: '($/MWh)' o '(USD$/MWh)'."""
    if get_moneda() == "USD":
        return "(USD$/MWh)"
    return "($/MWh)"


def label_moneda_short():
    """Retorna label corto sin /MWh."""
    if get_moneda() == "USD":
        return "USD$"
    return "$"

C_HEADER = "0e346b"
C_SUB    = "2777bd"
C_RED    = "a0090c"
C_WHITE  = "FFFFFF"
C_ALT    = "EBF3FB"
C_INFO   = "D9E2F3"
C_GREEN  = "1a8a3a"
C_GOLD   = "d4a017"

st.markdown(f"""
<style>
    .main-header {{
        background: linear-gradient(135deg, {RE_NAVY} 0%, {RE_BLUE} 100%);
        padding: 2rem; border-radius: 8px; margin-bottom: 2rem; color: white;
    }}
    .main-header h1 {{ color: white !important; margin: 0; font-size: 2rem; }}
    .main-header p {{ color: rgba(255,255,255,0.85); margin: 0.5rem 0 0 0; }}
    .srf-badge {{
        background: {RE_RED}; color: white; padding: 0.3rem 0.8rem;
        border-radius: 4px; font-weight: bold; font-size: 0.85rem;
        display: inline-block; margin-left: 1rem;
    }}
    .stButton > button[kind="primary"] {{
        background: {RE_NAVY}; color: white; border: none;
    }}
    .stButton > button[kind="primary"]:hover {{ background: {RE_RED}; }}
    .mode-badge {{
        background: #f0f4f8; border-left: 4px solid {RE_NAVY};
        padding: 0.6rem 1rem; border-radius: 4px; margin-bottom: 1rem;
        color: {TEXT_DARK}; font-size: 0.9rem;
    }}
    .ccr-legend {{
        background: white; border: 1px solid #d0d0d0;
        border-radius: 6px; padding: 0.8rem 1rem;
        margin-top: 0.5rem;
    }}
    .ccr-legend-item {{
        display: inline-block; margin: 0.2rem 0.6rem 0.2rem 0;
        font-size: 0.88rem; color: {TEXT_DARK};
    }}
    .ccr-dot {{
        display: inline-block; width: 12px; height: 12px;
        border-radius: 50%; margin-right: 5px;
        border: 1px solid {TEXT_DARK};
        vertical-align: middle;
    }}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
# CONSTANTES API
# ═══════════════════════════════════════════════════════════════════════
BASE_URL    = "https://ws01.cenace.gob.mx:8082/SWPML/SIM"
BLOQUE_MAX  = 7
FORMATO     = "JSON"
TIMEOUT     = 60

OSM_HEADERS = {
    'Accept': 'application/json',
    'Content-Type': 'application/x-www-form-urlencoded',
    'User-Agent': 'CENACE-PML-Analyzer/1.0 (Recurrent Energy MX, SRF)',
    'Referer': 'https://recurrent-energy.com/',
}


# ═══════════════════════════════════════════════════════════════════════
# CARGAR CATÁLOGO (cached)
# ═══════════════════════════════════════════════════════════════════════
@st.cache_data
def cargar_catalogo_default():
    """Carga el catálogo del repo (default)."""
    catalog_path = os.path.join(os.path.dirname(__file__), 'data', 'catalogo_nodos.xlsx')
    if not os.path.exists(catalog_path):
        return {}, None
    return _parse_catalogo_workbook(catalog_path), "repositorio (data/catalogo_nodos.xlsx)"


def cargar_catalogo_uploaded(file_bytes):
    """Carga catálogo desde bytes subidos por el usuario."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        catalogo = {}
        for r in ws.iter_rows(min_row=3, values_only=True):
            if not r or len(r) < 6 or not r[3]:
                continue
            clave = str(r[3]).strip()
            catalogo[clave] = {
                'sistema':   str(r[0]).strip() if r[0] else '',
                'ccr':       str(r[1]).strip() if r[1] else '',
                'zona':      str(r[2]).strip() if r[2] else '',
                'nombre':    str(r[4]).strip() if r[4] else '',
                'kv':        int(r[5]) if r[5] else 0,
                'estado':    str(r[15]).strip() if len(r) > 15 and r[15] else '',
                'municipio': str(r[17]).strip() if len(r) > 17 and r[17] else '',
            }
        wb.close()
        return catalogo
    except Exception as e:
        return {}


def _parse_catalogo_workbook(path):
    """Parser interno que convierte un xlsx a dict."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    catalogo = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 6 or not r[3]:
            continue
        clave = str(r[3]).strip()
        catalogo[clave] = {
            'sistema':   str(r[0]).strip() if r[0] else '',
            'ccr':       str(r[1]).strip() if r[1] else '',
            'zona':      str(r[2]).strip() if r[2] else '',
            'nombre':    str(r[4]).strip() if r[4] else '',
            'kv':        int(r[5]) if r[5] else 0,
            'estado':    str(r[15]).strip() if len(r) > 15 and r[15] else '',
            'municipio': str(r[17]).strip() if len(r) > 17 and r[17] else '',
        }
    wb.close()
    return catalogo


def get_catalogo_activo():
    """Retorna (catalogo, fuente) — catálogo subido por usuario o default del repo."""
    if "catalogo_uploaded" in st.session_state and st.session_state["catalogo_uploaded"]:
        return st.session_state["catalogo_uploaded"], "subido por usuario"
    cat, fuente = cargar_catalogo_default()
    return cat, fuente


# ═══════════════════════════════════════════════════════════════════════
# OSM CACHE
# ═══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False, ttl=86400 * 7)
def cargar_osm_subestaciones():
    osm_local = os.path.join(os.path.dirname(__file__), 'data', 'osm_subestaciones_mx.json')
    if os.path.exists(osm_local):
        with open(osm_local, 'r', encoding='utf-8') as f:
            return json.load(f)

    query = """
[out:json][timeout:180];
area["ISO3166-1"="MX"]->.mx;
(
  node["power"="substation"](area.mx);
  way["power"="substation"](area.mx);
);
out center tags;
"""
    try:
        r = requests.post('https://overpass-api.de/api/interpreter',
                         data={'data': query},
                         headers=OSM_HEADERS, timeout=200)
        if r.status_code != 200:
            return []
        data = r.json()
        elementos = data.get('elements', [])
    except Exception:
        return []

    def _parse_voltaje_max(v_raw):
        if not v_raw: return None
        v_clean = str(v_raw).replace('kV', '').replace('KV', '').replace('kv', '').strip()
        partes = [p.strip() for p in v_clean.split(';') if p.strip()]
        nums = []
        for p in partes:
            num_str = ''
            for ch in p:
                if ch.isdigit() or ch == '.':
                    num_str += ch
                else: break
            if num_str:
                try:
                    n = float(num_str)
                    nums.append(n / 1000 if n >= 1000 else n)
                except: pass
        return max(nums) if nums else None

    subs = []
    for el in elementos:
        tags = el.get('tags', {})
        sub_type = tags.get('substation', '')
        if sub_type in ('minor_distribution', 'traction', 'industrial',
                       'household', 'measurement'):
            continue
        if el['type'] == 'node':
            lat_e = el.get('lat', 0); lon_e = el.get('lon', 0)
        elif 'center' in el:
            lat_e = el['center'].get('lat', 0); lon_e = el['center'].get('lon', 0)
        else: continue
        if lat_e == 0 and lon_e == 0: continue
        v_kv = _parse_voltaje_max(tags.get('voltage', ''))
        subs.append({
            'osm_id': el['id'], 'osm_type': el['type'],
            'lat': round(lat_e, 6), 'lon': round(lon_e, 6),
            'name': tags.get('name', ''),
            'operator': tags.get('operator', ''),
            'voltage_kv': v_kv,
            'voltage_raw': tags.get('voltage', ''),
            'substation': sub_type,
            'rating': tags.get('rating', ''),
        })
    return subs


# ═══════════════════════════════════════════════════════════════════════
# FX — TIPO DE CAMBIO BANXICO (FIX serie SF43718)
# ═══════════════════════════════════════════════════════════════════════
BANXICO_API_BASE = "https://www.banxico.org.mx/SieAPIRest/service/v1/series"
SERIE_FIX_USD = "SF43718"  # Tipo de cambio FIX peso-dólar


@st.cache_data(show_spinner=False, ttl=86400)  # 24h cache
def cargar_fx_banxico(fecha_ini_yyyymmdd, fecha_fin_yyyymmdd, _token):
    """Descarga serie histórica del FIX entre dos fechas.

    Retorna dict {fecha_str: tipo_cambio_float} o {} si falla.
    Las fechas son en formato YYYY-MM-DD.
    """
    if not _token:
        return {}

    url = (f"{BANXICO_API_BASE}/{SERIE_FIX_USD}/datos/"
           f"{fecha_ini_yyyymmdd}/{fecha_fin_yyyymmdd}"
           f"?mediaType=json")
    headers = {'Bmx-Token': _token, 'Accept': 'application/json'}

    try:
        r = requests.get(url, headers=headers, timeout=30)
        if r.status_code != 200:
            return {}
        data = r.json()
        series = data.get('bmx', {}).get('series', [])
        if not series:
            return {}
        datos = series[0].get('datos', [])
    except Exception:
        return {}

    fx_dict = {}
    for d in datos:
        fecha_raw = d.get('fecha', '')  # formato dd/mm/yyyy
        valor_raw = d.get('dato', '')
        if not fecha_raw or not valor_raw or valor_raw == 'N/E':
            continue
        try:
            # convertir dd/mm/yyyy -> yyyy-mm-dd
            day, month, year = fecha_raw.split('/')
            fecha_norm = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
            fx_dict[fecha_norm] = float(valor_raw)
        except (ValueError, AttributeError):
            continue
    return fx_dict


def construir_fx_lookup(fx_dict, fechas_necesarias):
    """Para fechas sin TC (fines de semana, feriados), usa el último día hábil disponible.

    Retorna dict {fecha_str: tc_aplicable}.
    """
    if not fx_dict:
        return {}

    # Ordenar fechas disponibles
    fechas_disponibles = sorted(fx_dict.keys())
    if not fechas_disponibles:
        return {}

    lookup = {}
    fechas_solicitadas = sorted(set(fechas_necesarias))

    idx = 0  # puntero a fechas disponibles
    last_known = fx_dict[fechas_disponibles[0]]

    for fecha_req in fechas_solicitadas:
        # Avanzar puntero hasta encontrar fecha <= fecha_req
        while idx < len(fechas_disponibles) and fechas_disponibles[idx] <= fecha_req:
            last_known = fx_dict[fechas_disponibles[idx]]
            idx += 1
        lookup[fecha_req] = last_known
    return lookup


def aplicar_conversion_usd(acumulado, fx_lookup):
    """Aplica conversión MXN→USD a todos los registros de acumulado.

    Modifica `acumulado` in-place dividiendo PML, PML_ENE, PML_PER, PML_CNG
    entre el tipo de cambio del día. Retorna el acumulado modificado.
    """
    if not fx_lookup:
        return acumulado

    for nodo, filas in acumulado.items():
        for f in filas:
            fecha = f.get("fecha", "")
            tc = fx_lookup.get(fecha)
            if not tc or tc <= 0:
                continue
            for k in ("pml", "pml_ene", "pml_per", "pml_cng"):
                v = f.get(k)
                if isinstance(v, (int, float)):
                    f[k] = round(v / tc, 4)
    return acumulado


def obtener_token_banxico():
    """Lee el token de Streamlit secrets. Retorna '' si no existe."""
    try:
        return st.secrets.get("BANXICO_TOKEN", "")
    except Exception:
        return ""


# ═══════════════════════════════════════════════════════════════════════
# MATCHING GEOGRÁFICO
# ═══════════════════════════════════════════════════════════════════════
BBOX_ESTADO = {
    'AGUASCALIENTES':       (21.62, -102.87, 22.45, -101.84),
    'BAJA CALIFORNIA':      (28.00, -117.30, 32.72, -112.75),
    'BAJA CALIFORNIA SUR':  (22.87, -115.30, 28.00, -109.40),
    'CAMPECHE':             (17.81, -92.47,  20.85, -89.10),
    'CHIAPAS':              (14.53, -94.14,  17.99, -90.37),
    'CHIHUAHUA':            (25.55, -109.07, 31.78, -103.30),
    'COAHUILA':             (24.55, -103.96, 29.88, -99.84),
    'COLIMA':               (18.69, -104.74, 19.52, -103.49),
    'DURANGO':              (22.36, -107.20, 26.83, -102.46),
    'GUANAJUATO':           (19.91, -102.10, 21.85, -99.69),
    'GUERRERO':             (16.31, -102.18, 18.92, -98.00),
    'HIDALGO':              (19.63, -99.91,  21.40, -97.97),
    'JALISCO':              (18.92, -105.69, 22.75, -101.50),
    'MEXICO':               (18.36, -100.59, 20.29, -98.62),
    'CIUDAD DE MEXICO':     (19.05, -99.36,  19.59, -98.94),
    'MICHOACAN':            (17.92, -103.74, 20.40, -100.07),
    'MORELOS':              (18.32, -99.50,  19.13, -98.62),
    'NAYARIT':              (20.60, -105.78, 23.09, -103.71),
    'NUEVO LEON':           (23.18, -101.21, 27.81, -98.40),
    'OAXACA':               (15.65, -98.51,  18.66, -93.86),
    'PUEBLA':               (17.86, -99.06,  20.84, -96.71),
    'QUERETARO':            (20.02, -100.59, 21.67, -99.04),
    'QUINTANA ROO':         (17.88, -89.30,  21.62, -86.69),
    'SAN LUIS POTOSI':      (21.16, -102.30, 24.55, -98.34),
    'SINALOA':              (22.50, -109.45, 27.04, -105.41),
    'SONORA':               (26.32, -115.07, 32.51, -108.40),
    'TABASCO':              (17.25, -94.13,  18.65, -91.00),
    'TAMAULIPAS':           (22.21, -100.16, 27.68, -97.13),
    'TLAXCALA':             (19.06, -98.71,  19.72, -97.62),
    'VERACRUZ':             (17.15, -98.65,  22.46, -93.61),
    'YUCATAN':              (19.55, -90.42,  21.62, -87.53),
    'ZACATECAS':            (21.04, -104.34, 25.13, -100.81),
}

PALABRAS_GENERICAS = {
    'maniobras', 'planta', 'entronque', 'subestacion', 'subestación',
    'central', 'switching', 'rectificadora', 'tap', 'industrial',
    'cementos', 'pemex', 'cuadro', 'parque', 'potencia', 'aeropuerto',
    'norte', 'sur', 'este', 'oeste', 'poniente', 'oriente',
    'centro', 'cerro', 'valle', 'laguna', 'puerto', 'villa',
    'cabo', 'monte', 'sierra', 'mesa', 'isla', 'rio',
    'nuevo', 'nueva', 'viejo', 'vieja', 'gran', 'grande',
    'antiguo', 'antigua', 'segundo', 'segunda', 'primero', 'primera',
    'i', 'ii', 'iii', 'iv', 'v', 'vi', 'vii', 'viii', 'ix', 'x',
    '1', '2', '3', '4', '5', '6', '7', '8', '9', '10',
    'uno', 'dos', 'tres', 'cuatro', 'cinco',
    'de', 'la', 'el', 'las', 'los', 'del', 'al', 'con', 'por',
    'mexico', 'santa', 'santo', 'san',
}

PALABRAS_SEMIGENERICAS = {
    'hermosillo', 'guadalajara', 'monterrey', 'merida', 'chihuahua',
    'culiacan', 'puebla', 'queretaro', 'leon', 'morelia', 'tepic',
    'durango', 'mazatlan', 'oaxaca', 'campeche', 'cancun',
    'chetumal', 'colima', 'manzanillo', 'veracruz', 'tampico',
    'reynosa', 'matamoros', 'juarez', 'mexicali', 'tijuana',
    'ensenada', 'cardenas', 'guadalupe', 'progreso', 'lazaro',
}


def en_estado(lat, lon, nombre_estado):
    if not nombre_estado: return None
    clave = nombre_estado.upper().strip()
    for variant in [clave, clave.split(' ')[0],
                    ' '.join(clave.split(' ')[:2]),
                    ' '.join(clave.split(' ')[:3])]:
        bb = BBOX_ESTADO.get(variant)
        if bb:
            return bb[0] <= lat <= bb[2] and bb[1] <= lon <= bb[3]
    return None


def normalizar_nombre(s):
    if not s: return ''
    s = s.lower().strip()
    repl = {'á':'a','é':'e','í':'i','ó':'o','ú':'u','ñ':'n','ü':'u'}
    for k, v in repl.items():
        s = s.replace(k, v)
    for prefix in ['s/e ', 's/e. ', 'se ', 'sub ', 'subestacion ',
                   'central ', 'planta ', 'c.t. ', 'ct ',
                   'c.h. ', 'ch ', 'c.f.e. ', 'cfe ']:
        if s.startswith(prefix):
            s = s[len(prefix):]
    for art in ['los ', 'las ', 'el ', 'la ']:
        if s.startswith(art):
            resto = s[len(art):]
            if len(resto.split(' ', 1)[0]) >= 4:
                s = resto
                break
    for sufijo in [' maniobras', ' switching', ' tap',
                   ' planta', ' central', ' cfe']:
        if s.endswith(sufijo):
            s = s[:-len(sufijo)]
    roman_map = {' i': ' 1', ' ii': ' 2', ' iii': ' 3', ' iv': ' 4',
                 ' v': ' 5', ' vi': ' 6', ' vii': ' 7', ' viii': ' 8',
                 ' ix': ' 9', ' x': ' 10'}
    for r, a in roman_map.items():
        if s.endswith(r):
            s = s[:-len(r)] + a
            break
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def similitud(a, b):
    if not a or not b: return 0.0
    return SequenceMatcher(None, a, b).ratio()


def palabras_clave(nombre_norm):
    if not nombre_norm: return []
    return [p for p in nombre_norm.split()
            if len(p) >= 5 and p not in PALABRAS_GENERICAS and not p.isdigit()]


def buscar_match(nombre_cenace, voltaje_cenace_kv, estado_cat, candidatos):
    nombre_norm = normalizar_nombre(nombre_cenace)
    if not nombre_norm:
        return None, 0.0, 'sin nombre'

    exactos = [s for s in candidatos if s['name_norm'] == nombre_norm]
    if exactos:
        if estado_cat:
            exactos_en_est = [s for s in exactos
                              if en_estado(s['lat'], s['lon'], estado_cat) is not False]
            if exactos_en_est:
                exactos = exactos_en_est
            elif len(nombre_norm) < 18:
                exactos = []
        if exactos:
            if voltaje_cenace_kv > 0 and len(exactos) > 1:
                con_v = [s for s in exactos if s['voltage_kv'] and
                         abs(s['voltage_kv'] - voltaje_cenace_kv) < voltaje_cenace_kv * 0.15]
                if con_v:
                    return con_v[0], 1.0, 'match exacto + voltaje'
            return exactos[0], 1.0, 'match exacto'

    substring_matches = []
    for s in candidatos:
        n_osm = s['name_norm']
        if nombre_norm in n_osm and len(nombre_norm) >= 4:
            if n_osm.startswith(nombre_norm + ' ') or n_osm == nombre_norm:
                cob = len(nombre_norm) / len(n_osm)
                if cob >= 0.40:
                    substring_matches.append((0.92, s, 'OSM extiende CENACE'))
            elif (' ' + nombre_norm) in n_osm or n_osm.endswith(' ' + nombre_norm):
                cob = len(nombre_norm) / len(n_osm)
                if cob >= 0.40:
                    substring_matches.append((0.88, s, 'CENACE dentro de OSM'))
        elif n_osm in nombre_norm and len(n_osm) >= 4:
            if nombre_norm.startswith(n_osm + ' ') or nombre_norm == n_osm:
                cob = len(n_osm) / len(nombre_norm)
                if cob >= 0.40:
                    substring_matches.append((0.90, s, 'CENACE extiende OSM'))

    difusos = []
    for s in candidatos:
        sim = similitud(nombre_norm, s['name_norm'])
        if sim >= 0.80:
            difusos.append((sim, s))

    kw_cenace = palabras_clave(nombre_norm)
    keyword_matches = []
    if kw_cenace:
        for s in candidatos:
            kw_osm = palabras_clave(s.get('name_norm', ''))
            if not kw_osm: continue
            compartidas = set(kw_cenace) & set(kw_osm)
            if not compartidas: continue
            todas_semi = all(p in PALABRAS_SEMIGENERICAS for p in compartidas)
            largas_no_semi = sum(1 for p in compartidas
                                 if len(p) >= 6 and p not in PALABRAS_SEMIGENERICAS)
            if todas_semi: base_score = 0.62
            elif largas_no_semi >= 1: base_score = 0.78
            else: base_score = 0.70
            ratio = len(compartidas) / max((len(kw_cenace) + len(kw_osm)) / 2, 1)
            score = base_score + (ratio * 0.12)
            keyword_matches.append((score, s, list(compartidas)))

    scored = []; seen = set()
    for sc, s, t in substring_matches:
        if s['osm_id'] not in seen:
            scored.append((sc, s, 'substring', t))
            seen.add(s['osm_id'])
    for sim, s in difusos:
        if s['osm_id'] not in seen:
            scored.append((sim, s, 'difuso', ''))
            seen.add(s['osm_id'])
    for sc, s, kws in keyword_matches:
        if s['osm_id'] not in seen:
            scored.append((sc, s, 'keyword', f'comparte "{",".join(kws[:2])}"'))
            seen.add(s['osm_id'])

    if not scored:
        return None, 0.0, 'no match'

    if estado_cat:
        scored = [(sc, s, t, st) for sc, s, t, st in scored
                  if en_estado(s['lat'], s['lon'], estado_cat) is not False]

    if not scored:
        return None, 0.0, 'sin match (estado)'

    if voltaje_cenace_kv > 0:
        filt = []
        for sim, s, t, st in scored:
            v = s['voltage_kv']
            ok = (v is not None and abs(v - voltaje_cenace_kv) < voltaje_cenace_kv * 0.15)
            unk = v is None
            if t == 'difuso' and sim < 0.92 and not (ok or unk):
                continue
            if t == 'keyword' and not (ok or unk):
                continue
            if ok: filt.append((sim + 0.02, s, t, st))
            elif unk: filt.append((sim, s, t, st))
            else: filt.append((sim - 0.10, s, t, st))
        scored = filt

    if not scored:
        return None, 0.0, 'sin match (voltaje)'

    scored.sort(key=lambda x: -x[0])
    sim_f, mejor, tipo, sub_tipo = scored[0]
    if tipo == 'substring': razon = sub_tipo
    elif tipo == 'keyword': razon = f'palabra-clave: {sub_tipo}'
    elif sim_f >= 0.95: razon = 'match difuso alto'
    elif sim_f >= 0.85: razon = 'match difuso medio'
    else: razon = f'match difuso ({sim_f:.2f})'
    return mejor, sim_f, razon


def matchear_nodos(nodos, catalogo, osm_subs):
    for s in osm_subs:
        if 'name_norm' not in s:
            s['name_norm'] = normalizar_nombre(s.get('name', ''))
    candidatos = [s for s in osm_subs if s.get('name_norm')]
    resultados = []
    for clave in nodos:
        info = catalogo.get(clave, {})
        nombre = info.get('nombre', clave)
        kv = info.get('kv', 0)
        if kv == 0:
            try: kv = int(clave.split('-')[-1])
            except: kv = 0
        ccr = info.get('ccr', '')
        zona = info.get('zona', '')
        estado = info.get('estado', '')
        municipio = info.get('municipio', '')

        match, sim, razon = buscar_match(nombre, kv, estado, candidatos)

        if match:
            if 'palabra-clave' in razon: calidad = '🥉 Aceptable'
            elif sim >= 0.95: calidad = '🥇 Excelente'
            elif sim >= 0.90: calidad = '🥈 Bueno'
            else: calidad = '🥉 Aceptable'
        else:
            calidad = '❌ Sin match'

        resultados.append({
            'clave': clave, 'nombre_cenace': nombre, 'ccr': ccr, 'zona': zona,
            'kv_cenace': kv, 'estado': estado, 'municipio': municipio,
            'calidad': calidad,
            'similitud': round(sim, 3) if match else 0,
            'razon': razon,
            'nombre_osm': match['name'] if match else '',
            'kv_osm': match['voltage_kv'] if match and match['voltage_kv'] else None,
            'voltage_raw': match['voltage_raw'] if match else '',
            'operator_osm': match['operator'] if match else '',
            'subtipo_osm': match['substation'] if match else '',
            'lat': match['lat'] if match else None,
            'lon': match['lon'] if match else None,
            'osm_id': match['osm_id'] if match else None,
            'osm_type': match['osm_type'] if match else None,
        })
    return resultados


# ═══════════════════════════════════════════════════════════════════════
# CENACE — descarga
# ═══════════════════════════════════════════════════════════════════════
def parse_fecha(s): return datetime.strptime(s, "%Y/%m/%d")
def fmt(d): return d.strftime("%Y/%m/%d")

def generar_bloques(ini_str, fin_str, max_dias=BLOQUE_MAX):
    ini = parse_fecha(ini_str); fin = parse_fecha(fin_str)
    bloques = []; cursor = ini
    while cursor <= fin:
        bloque_fin = min(cursor + timedelta(days=max_dias - 1), fin)
        bloques.append((fmt(cursor), fmt(bloque_fin)))
        cursor = bloque_fin + timedelta(days=1)
    return bloques

def construir_url(nodos_lista, fecha_ini, fecha_fin, sistema, proceso):
    nodos_str = ",".join(nodos_lista)
    ai, mi, di = fecha_ini.split("/")
    af, mf, df_ = fecha_fin.split("/")
    return f"{BASE_URL}/{sistema}/{proceso}/{nodos_str}/{ai}/{mi}/{di}/{af}/{mf}/{df_}/{FORMATO}"

def consultar(nodos_lista, fecha_ini, fecha_fin, sistema, proceso,
              errores_lista, lock, max_reintentos=4):
    url = construir_url(nodos_lista, fecha_ini, fecha_fin, sistema, proceso)
    if len(url) > 2000:
        with lock:
            errores_lista.append({"fecha": f"{fecha_ini} → {fecha_fin}",
                                   "nodos": len(nodos_lista),
                                   "error": "URL muy larga"})
        return None
    session = requests.Session(); session.verify = False
    for intento in range(max_reintentos):
        try:
            r = session.get(url, timeout=TIMEOUT)
            if r.status_code == 200: return r.text
            elif r.status_code == 204: return None
            elif r.status_code == 429:
                time.sleep(2 ** (intento + 2)); continue
            else:
                if intento < max_reintentos - 1:
                    time.sleep(2 ** (intento + 1)); continue
                with lock:
                    errores_lista.append({"fecha": f"{fecha_ini} → {fecha_fin}",
                                           "nodos": len(nodos_lista),
                                           "error": f"HTTP {r.status_code}"})
                return None
        except requests.exceptions.Timeout:
            if intento < max_reintentos - 1:
                time.sleep(2 ** intento); continue
            with lock:
                errores_lista.append({"fecha": f"{fecha_ini} → {fecha_fin}",
                                       "nodos": len(nodos_lista),
                                       "error": "Timeout"})
            return None
        except Exception as e:
            if intento < max_reintentos - 1:
                time.sleep(2 ** intento); continue
            with lock:
                errores_lista.append({"fecha": f"{fecha_ini} → {fecha_fin}",
                                       "nodos": len(nodos_lista),
                                       "error": f"{type(e).__name__}"})
            return None
    return None

def _num(val):
    try: return float(val)
    except (TypeError, ValueError): return val

def parsear_json(texto):
    if not texto: return {}
    try: obj = json.loads(texto)
    except Exception: return {}
    datos = {}
    for nd in obj.get("Resultados", []):
        clv = nd.get("clv_nodo", "?")
        valores = nd.get("Valores", [])
        if not valores: continue
        registros = []
        for v in valores:
            registros.append({
                "fecha":   v.get("fecha", ""),
                "hora":    int(v.get("hora", 0)) if v.get("hora") else 0,
                "pml":     _num(v.get("pml", 0)),
                "pml_ene": _num(v.get("pml_ene", 0)),
                "pml_per": _num(v.get("pml_per", 0)),
                "pml_cng": _num(v.get("pml_cng", 0)),
            })
        datos[clv] = registros
    return datos

def descargar_pml(nodos, fecha_ini, fecha_fin, sistema, proceso, progress_cb=None):
    bloques = generar_bloques(fecha_ini, fecha_fin)
    LOTE = 10
    lotes_nodos = [nodos[i:i+LOTE] for i in range(0, len(nodos), LOTE)]
    consultas = []
    for lote in lotes_nodos:
        for bi, bf in bloques:
            consultas.append((lote, bi, bf))
    total = len(consultas)
    acumulado = {n: [] for n in nodos}
    errores_consulta = []; lock = Lock(); completed = 0
    def _job(c):
        lote, bi, bf = c
        return parsear_json(consultar(lote, bi, bf, sistema, proceso,
                                       errores_consulta, lock))
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_job, c): c for c in consultas}
        for fut in as_completed(futures):
            data = fut.result() or {}
            for nodo, regs in data.items():
                if nodo in acumulado:
                    acumulado[nodo].extend(regs)
            completed += 1
            if progress_cb:
                progress_cb(completed, total)
    acumulado = {n: regs for n, regs in acumulado.items() if regs}
    return acumulado, errores_consulta


def descargar_pml_auto(nodos, fecha_ini, fecha_fin, proceso, catalogo, progress_cb=None):
    """Auto-detecta el Sistema de cada nodo desde el catálogo y descarga.

    Si hay nodos de varios sistemas (raro: BCA/BCS son redes aisladas), hace
    consultas paralelas y combina resultados.

    Retorna (acumulado, errores, info_sistemas) donde info_sistemas es un dict
    {sistema: count_nodos}.
    """
    # Agrupar nodos por sistema según catálogo
    grupos = {"SIN": [], "BCA": [], "BCS": []}
    nodos_sin_sistema = []

    for nodo in nodos:
        info = catalogo.get(nodo, {}) if catalogo else {}
        sistema = (info.get("sistema") or "").strip().upper()
        if sistema in grupos:
            grupos[sistema].append(nodo)
        else:
            # Fallback: asumir SIN (95% de los nodos están ahí)
            nodos_sin_sistema.append(nodo)
            grupos["SIN"].append(nodo)

    info_sistemas = {s: len(ns) for s, ns in grupos.items() if ns}

    # Si todo es un solo sistema, llamada directa (sin overhead)
    sistemas_activos = [s for s, ns in grupos.items() if ns]
    if len(sistemas_activos) == 1:
        sistema = sistemas_activos[0]
        acumulado, errores = descargar_pml(
            grupos[sistema], fecha_ini, fecha_fin, sistema, proceso, progress_cb=progress_cb)
        return acumulado, errores, info_sistemas, nodos_sin_sistema

    # Multi-sistema: descargar cada uno en serie (no paralelo entre sistemas
    # para no saturar el progress bar; cada sistema YA usa workers paralelos internos)
    acumulado_total = {}
    errores_total = []

    # Progress bar combinado
    completed_jobs = [0]
    total_jobs_estimate = sum(
        len(generar_bloques(fecha_ini, fecha_fin)) * ((len(grupos[s]) + 9) // 10)
        for s in sistemas_activos
    )

    def cb_combined(done, total_per_system):
        if progress_cb and total_jobs_estimate > 0:
            current = completed_jobs[0] + done
            progress_cb(current, total_jobs_estimate)

    for sistema in sistemas_activos:
        nodos_grupo = grupos[sistema]
        if not nodos_grupo:
            continue
        n_jobs_grupo = len(generar_bloques(fecha_ini, fecha_fin)) * ((len(nodos_grupo) + 9) // 10)
        ac_g, err_g = descargar_pml(nodos_grupo, fecha_ini, fecha_fin, sistema, proceso,
                                      progress_cb=cb_combined)
        acumulado_total.update(ac_g)
        errores_total.extend(err_g)
        completed_jobs[0] += n_jobs_grupo

    return acumulado_total, errores_total, info_sistemas, nodos_sin_sistema


# ═══════════════════════════════════════════════════════════════════════
# EXCEL DATOS (datos crudos)
# ═══════════════════════════════════════════════════════════════════════
def generar_excel_datos(acumulado, sistema, proceso, fecha_ini, fecha_fin,
                          moneda="MXN", tc_info=None):
    _NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
    _side = Side(style="thin", color="BFBFBF")
    BORDE = Border(left=_side, right=_side, top=_side, bottom=_side)

    _sym = "USD$" if moneda == "USD" else "$"

    def hdr(cell, bg=C_HEADER, fg=C_WHITE, size=11):
        cell.font = Font(bold=True, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE

    _FONT_DATO = Font(name="Arial", size=10)
    _ALIGN_NUM = Alignment(horizontal="center", vertical="center")
    _ALIGN_TXT = Alignment(horizontal="left", vertical="center")
    _FILL_ALT  = PatternFill("solid", start_color=C_ALT)

    COLS   = ["Fecha", "Hora", f"PML ({_sym}/MWh)", "Energía", "Pérdidas", "Congestión"]
    ANCHOS = [14, 8, 18, 16, 16, 18]
    ES_NUM = [False, True, True, True, True, True]

    wb = Workbook()
    ws = wb.active
    ws.title = "■ Portada"
    ws.sheet_properties.tabColor = C_HEADER
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H4")
    c = ws["B2"]
    c.value = "Precios Marginales Locales\nMercado Eléctrico Mayorista — CENACE"
    c.font = Font(bold=True, color=C_WHITE, size=22, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, 5):
        ws.row_dimensions[r].height = 32

    ws.merge_cells("B5:H5")
    c = ws["B5"]; c.value = "Reporte de Datos"
    c.font = Font(bold=True, italic=True, color=C_WHITE, size=12, name="Arial")
    c.fill = PatternFill("solid", start_color=C_RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24

    ws.merge_cells("B7:H7")
    c = ws["B7"]; c.value = "PARÁMETROS DE LA CONSULTA"
    c.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    c.fill = PatternFill("solid", start_color=C_SUB)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    info_rows = [
        ("Sistema", sistema), ("Proceso", proceso),
        ("Fecha inicial", fecha_ini), ("Fecha final", fecha_fin),
        ("Nodos en reporte", f"{len(acumulado):,} nodos"),
        ("Total registros", f"{sum(len(f) for f in acumulado.values()):,} filas"),
        ("Moneda", f"{moneda}" + (f" · TC FIX promedio: {tc_info['tc_promedio']:.4f}" if tc_info and tc_info.get('tc_promedio') else "")),
        ("Fecha generación", _NOW),
    ]
    for ri, (lbl, val) in enumerate(info_rows, start=8):
        ws.merge_cells(f"B{ri}:D{ri}")
        c1 = ws[f"B{ri}"]; c1.value = lbl
        c1.font = Font(bold=True, color=C_HEADER, size=11, name="Arial")
        c1.fill = PatternFill("solid", start_color=C_INFO)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = BORDE
        ws.merge_cells(f"E{ri}:H{ri}")
        c2 = ws[f"E{ri}"]; c2.value = val
        c2.font = Font(size=11, name="Arial")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border = BORDE
        ws.row_dimensions[ri].height = 22

    # Footer discreto con SRF
    ws.merge_cells("B17:H17")
    c = ws["B17"]; c.value = "Prepared by: Sebastian Roldan (SRF) · Recurrent Energy"
    c.font = Font(italic=True, bold=True, color=C_HEADER, size=12, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=C_INFO)
    ws.row_dimensions[17].height = 24

    ws.merge_cells("B18:H18")
    c = ws["B18"]; c.value = "A subsidiary of Canadian Solar"
    c.font = Font(italic=True, color="555555", size=10, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 18

    for col, w in [("A", 3), ("B", 14), ("C", 14), ("D", 14),
                   ("E", 14), ("F", 14), ("G", 14), ("H", 14), ("I", 3)]:
        ws.column_dimensions[col].width = w

    ws_res = wb.create_sheet("Resumen")
    ws_res.merge_cells("A1:F1")
    c = ws_res["A1"]; c.value = f"Resumen — {len(acumulado)} nodos · {sistema} · {proceso}"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 28

    res_cols = ["Nodo", "# Registros", "PML promedio", "PML máximo", "PML mínimo", "Período"]
    for ci, col in enumerate(res_cols, 1):
        hdr(ws_res.cell(row=2, column=ci, value=col), bg=C_SUB)
    ws_res.row_dimensions[2].height = 22

    for ri, (nodo, filas) in enumerate(acumulado.items(), start=3):
        if not filas: continue
        pmls = [f["pml"] for f in filas if isinstance(f["pml"], (int, float))]
        bg = C_ALT if ri % 2 == 0 else None
        valores = [
            nodo, len(filas),
            f"{sum(pmls)/len(pmls):.2f}" if pmls else "—",
            f"{max(pmls):.2f}" if pmls else "—",
            f"{min(pmls):.2f}" if pmls else "—",
            f"{filas[0]['fecha']} → {filas[-1]['fecha']}",
        ]
        for ci, v in enumerate(valores, 1):
            cc = ws_res.cell(row=ri, column=ci, value=v)
            cc.font = _FONT_DATO
            cc.alignment = _ALIGN_NUM if ci > 1 else _ALIGN_TXT
            cc.border = BORDE
            if bg: cc.fill = PatternFill("solid", start_color=bg)

    for ci, w in enumerate([14, 14, 14, 14, 14, 24], 1):
        ws_res.column_dimensions[get_column_letter(ci)].width = w
    ws_res.freeze_panes = "A3"

    for nodo, filas in acumulado.items():
        ws_n = wb.create_sheet(title=nodo[:31])
        ws_n.merge_cells("A1:F1")
        c = ws_n["A1"]; c.value = f"PML — Nodo: {nodo}"
        c.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_n.row_dimensions[1].height = 28

        ws_n.merge_cells("A2:F2")
        c = ws_n["A2"]
        c.value = (f"Sistema: {sistema} | Proceso: {proceso} | "
                   f"Período: {fecha_ini} → {fecha_fin} | Total: {len(filas):,} | "
                   f"Sebastian Roldan (SRF) · Recurrent Energy")
        c.font = Font(italic=True, size=9, name="Arial", color="555555")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=C_INFO)
        ws_n.row_dimensions[2].height = 16

        for ci, (enc, ancho) in enumerate(zip(COLS, ANCHOS), 1):
            hdr(ws_n.cell(row=3, column=ci, value=enc), bg=C_SUB)
            ws_n.column_dimensions[get_column_letter(ci)].width = ancho
        ws_n.row_dimensions[3].height = 22

        for fila in filas:
            ws_n.append([fila["fecha"], fila["hora"], fila["pml"],
                         fila["pml_ene"], fila["pml_per"], fila["pml_cng"]])
        for i in range(len(filas)):
            row_idx = i + 4
            es_alt = (i % 2 == 0)
            for ci in range(1, 7):
                cc = ws_n.cell(row=row_idx, column=ci)
                cc.font = _FONT_DATO
                cc.border = BORDE
                cc.alignment = _ALIGN_NUM if ES_NUM[ci-1] else _ALIGN_TXT
                if es_alt: cc.fill = _FILL_ALT
                if ES_NUM[ci-1]: cc.number_format = "#,##0.00"
        ws_n.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# EXCEL ANÁLISIS (BESS Scoring + métricas)
# ═══════════════════════════════════════════════════════════════════════
def generar_excel_analisis(df_metricas, df_resumen, sistema, proceso, fecha_ini, fecha_fin,
                             moneda="MXN", tc_info=None, df_multianos=None,
                             nodo_multianos=None):
    """Excel con BESS scoring para los 3 casos de uso + métricas + (opcional) gráfica multi-año.

    df_multianos: DataFrame con columnas mes_num + un col por año, con promedio PML.
    nodo_multianos: nombre del nodo para el cual se calculó df_multianos.
    """
    _NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
    _side = Side(style="thin", color="BFBFBF")
    BORDE = Border(left=_side, right=_side, top=_side, bottom=_side)

    _sym = "USD$" if moneda == "USD" else "$"

    def hdr(cell, bg=C_HEADER, fg=C_WHITE, size=11):
        cell.font = Font(bold=True, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE

    _FONT_DATO = Font(name="Arial", size=10)
    _ALIGN_NUM = Alignment(horizontal="center", vertical="center")
    _ALIGN_TXT = Alignment(horizontal="left", vertical="center")
    _FILL_ALT  = PatternFill("solid", start_color=C_ALT)

    wb = Workbook()
    ws = wb.active
    ws.title = "■ Portada"
    ws.sheet_properties.tabColor = C_RED
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H4")
    c = ws["B2"]
    c.value = "Análisis BESS — PML CENACE\nScoring por caso de uso"
    c.font = Font(bold=True, color=C_WHITE, size=22, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, 5):
        ws.row_dimensions[r].height = 32

    ws.merge_cells("B5:H5")
    c = ws["B5"]; c.value = "Reporte de Análisis — Battery Energy Storage System"
    c.font = Font(bold=True, italic=True, color=C_WHITE, size=12, name="Arial")
    c.fill = PatternFill("solid", start_color=C_RED)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 24

    ws.merge_cells("B7:H7")
    c = ws["B7"]; c.value = "PARÁMETROS"
    c.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
    c.fill = PatternFill("solid", start_color=C_SUB)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22

    info_rows = [
        ("Sistema", sistema), ("Proceso", proceso),
        ("Período", f"{fecha_ini} → {fecha_fin}"),
        ("Nodos analizados", f"{len(df_metricas)} nodos"),
        ("Casos de uso", "Arbitraje · SSAA · Renewables Firming"),
        ("Metodología scoring", "Rank-percentile (0=peor, 100=mejor)"),
        ("Moneda", f"{moneda}" + (f" · TC FIX promedio: {tc_info['tc_promedio']:.4f}" if tc_info and tc_info.get('tc_promedio') else "")),
        ("Fecha generación", _NOW),
    ]
    for ri, (lbl, val) in enumerate(info_rows, start=8):
        ws.merge_cells(f"B{ri}:D{ri}")
        c1 = ws[f"B{ri}"]; c1.value = lbl
        c1.font = Font(bold=True, color=C_HEADER, size=11, name="Arial")
        c1.fill = PatternFill("solid", start_color=C_INFO)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = BORDE
        ws.merge_cells(f"E{ri}:H{ri}")
        c2 = ws[f"E{ri}"]; c2.value = val
        c2.font = Font(size=11, name="Arial")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border = BORDE
        ws.row_dimensions[ri].height = 22

    # Footer discreto con SRF
    ws.merge_cells("B17:H17")
    c = ws["B17"]; c.value = "Prepared by: Sebastian Roldan (SRF) · Recurrent Energy"
    c.font = Font(italic=True, bold=True, color=C_HEADER, size=12, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=C_INFO)
    ws.row_dimensions[17].height = 24

    ws.merge_cells("B18:H18")
    c = ws["B18"]; c.value = "A subsidiary of Canadian Solar"
    c.font = Font(italic=True, color="555555", size=10, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[18].height = 18

    for col, w in [("A", 3), ("B", 16), ("C", 16), ("D", 16),
                   ("E", 16), ("F", 16), ("G", 16), ("H", 16), ("I", 3)]:
        ws.column_dimensions[col].width = w

    # Hoja Top 5 por caso de uso
    ws_top = wb.create_sheet("🏆 Top 5 por Caso de Uso")
    ws_top.merge_cells("A1:F1")
    c = ws_top["A1"]; c.value = "Top 5 nodos recomendados por caso de uso BESS"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_top.row_dimensions[1].height = 28

    cur_row = 3
    for use_case in ['Arbitraje', 'Servicios Auxiliares', 'Renewables Firming']:
        df_score = calcular_score_bess(df_metricas, use_case)
        if df_score.empty: continue

        # Header del caso
        ws_top.merge_cells(f"A{cur_row}:F{cur_row}")
        c = ws_top.cell(row=cur_row, column=1)
        c.value = f"⚡ {use_case}"
        c.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
        c.fill = PatternFill("solid", start_color=C_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_top.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Subheader
        cols_top = ["Rank", "Nodo", "Score", "PML promedio", "Volatilidad", "Spread P95-P5"]
        for ci, col in enumerate(cols_top, 1):
            hdr(ws_top.cell(row=cur_row, column=ci, value=col), bg=C_SUB)
        ws_top.row_dimensions[cur_row].height = 20
        cur_row += 1

        for i, (_, r) in enumerate(df_score.head(5).iterrows(), 1):
            bg = C_ALT if i % 2 == 0 else None
            medal = ['🥇', '🥈', '🥉', '4', '5'][i-1]
            valores = [medal, r['nodo'], f"{r['score']:.1f}",
                       f"${r['pml_promedio']:.2f}",
                       f"${r['volatilidad']:.2f}",
                       f"${r['spread_p95_p5']:.2f}"]
            for ci, v in enumerate(valores, 1):
                cc = ws_top.cell(row=cur_row, column=ci, value=v)
                cc.font = _FONT_DATO
                cc.border = BORDE
                cc.alignment = _ALIGN_NUM if ci != 2 else _ALIGN_TXT
                if bg: cc.fill = PatternFill("solid", start_color=bg)
            cur_row += 1
        cur_row += 1

    for ci, w in enumerate([8, 16, 14, 16, 16, 18], 1):
        ws_top.column_dimensions[get_column_letter(ci)].width = w
    ws_top.freeze_panes = "A2"

    # Hoja BESS scoring completo (3 casos de uso)
    for use_case in ['Arbitraje', 'Servicios Auxiliares', 'Renewables Firming']:
        df_score = calcular_score_bess(df_metricas, use_case)
        if df_score.empty: continue

        sheet_name = f"BESS {use_case[:22]}"
        ws_bess = wb.create_sheet(title=sheet_name[:31])

        ws_bess.merge_cells("A1:J1")
        c = ws_bess["A1"]; c.value = f"BESS Scoring — {use_case}"
        c.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_bess.row_dimensions[1].height = 26

        ws_bess.merge_cells("A2:J2")
        c = ws_bess["A2"]
        c.value = DESCRIPCIONES_USE_CASE_PLAIN.get(use_case, "")
        c.font = Font(italic=True, size=10, name="Arial", color="555555")
        c.fill = PatternFill("solid", start_color=C_INFO)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_bess.row_dimensions[2].height = 20

        cols = ["Rank", "Nodo", "Score (0-100)", "PML promedio",
                "Volatilidad", "Spread P95-P5", "Spread día prom",
                "Cambios bruscos", "Horas pico", "% horas neg"]
        for ci, col in enumerate(cols, 1):
            hdr(ws_bess.cell(row=3, column=ci, value=col), bg=C_SUB)
        ws_bess.row_dimensions[3].height = 32

        for i, (_, r) in enumerate(df_score.iterrows(), 1):
            bg = C_ALT if i % 2 == 0 else None
            valores = [i, r['nodo'], r['score'], r['pml_promedio'],
                       r['volatilidad'], r['spread_p95_p5'],
                       r['spread_avg_diario'], r['cambios_bruscos'],
                       r['horas_pico'], r['pct_horas_neg']]
            for ci, v in enumerate(valores, 1):
                cc = ws_bess.cell(row=i+3, column=ci, value=v)
                cc.font = _FONT_DATO
                cc.border = BORDE
                cc.alignment = _ALIGN_NUM if ci != 2 else _ALIGN_TXT
                if bg: cc.fill = PatternFill("solid", start_color=bg)
                if ci in (3, 4, 5, 6, 7): cc.number_format = "#,##0.00"
                elif ci == 10: cc.number_format = "0.0\"%\""

        for ci, w in enumerate([8, 16, 14, 14, 14, 14, 14, 14, 12, 14], 1):
            ws_bess.column_dimensions[get_column_letter(ci)].width = w
        ws_bess.freeze_panes = "A4"

    # Hoja Resumen (estadísticas básicas)
    ws_res = wb.create_sheet("📊 Resumen Estadístico")
    ws_res.merge_cells("A1:K1")
    c = ws_res["A1"]; c.value = f"Resumen estadístico — {len(df_resumen)} nodos"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 26

    cols = ["Nodo", "Nombre", "CCR", "# Reg", "Promedio", "Mediana",
            "Máximo", "Mínimo", "Volatilidad", "P95", "% horas neg"]
    for ci, col in enumerate(cols, 1):
        hdr(ws_res.cell(row=2, column=ci, value=col), bg=C_SUB)
    ws_res.row_dimensions[2].height = 22

    for i, (_, r) in enumerate(df_resumen.iterrows(), 1):
        bg = C_ALT if i % 2 == 0 else None
        valores = [r['nodo'], r['nombre'], r['ccr'], r['registros'],
                   r['promedio'], r['mediana'], r['maximo'], r['minimo'],
                   r['std'], r['p95'], r['% horas neg']]
        for ci, v in enumerate(valores, 1):
            cc = ws_res.cell(row=i+2, column=ci, value=v)
            cc.font = _FONT_DATO
            cc.border = BORDE
            cc.alignment = _ALIGN_NUM if ci > 3 else _ALIGN_TXT
            if bg: cc.fill = PatternFill("solid", start_color=bg)
            if ci in (5, 6, 7, 8, 9, 10): cc.number_format = "#,##0.00"
            elif ci == 11: cc.number_format = "0.0\"%\""

    for ci, w in enumerate([14, 22, 14, 10, 12, 12, 12, 12, 12, 12, 12], 1):
        ws_res.column_dimensions[get_column_letter(ci)].width = w
    ws_res.freeze_panes = "A3"

    # Hoja Métricas Raw (todas las métricas calculadas)
    ws_m = wb.create_sheet("🔢 Métricas Calculadas")
    ws_m.merge_cells("A1:I1")
    c = ws_m["A1"]; c.value = "Métricas BESS calculadas (datos raw)"
    c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws_m.row_dimensions[1].height = 26

    cols_m = ["Nodo", "PML promedio", "Volatilidad", "Spread P95-P5",
              "Spread día prom", "Spread día/noche", "Cambios bruscos",
              "Horas pico", "% horas neg"]
    for ci, col in enumerate(cols_m, 1):
        hdr(ws_m.cell(row=2, column=ci, value=col), bg=C_SUB)
    ws_m.row_dimensions[2].height = 32

    for i, (_, r) in enumerate(df_metricas.iterrows(), 1):
        bg = C_ALT if i % 2 == 0 else None
        valores = [r['nodo'], r['pml_promedio'], r['volatilidad'],
                   r['spread_p95_p5'], r['spread_avg_diario'], r['spread_dia'],
                   r['cambios_bruscos'], r['horas_pico'], r['pct_horas_neg']]
        for ci, v in enumerate(valores, 1):
            cc = ws_m.cell(row=i+2, column=ci, value=v)
            cc.font = _FONT_DATO
            cc.border = BORDE
            cc.alignment = _ALIGN_NUM if ci != 1 else _ALIGN_TXT
            if bg: cc.fill = PatternFill("solid", start_color=bg)
            if ci in (2, 3, 4, 5, 6): cc.number_format = "#,##0.00"
            elif ci == 9: cc.number_format = "0.0\"%\""

    for ci, w in enumerate([14, 14, 14, 14, 14, 14, 14, 12, 12], 1):
        ws_m.column_dimensions[get_column_letter(ci)].width = w
    ws_m.freeze_panes = "A3"

    # ─── HOJAS OPCIONALES: gráficas multi-año (1 por nodo) ───
    # df_multianos puede ser:
    #   (a) Un DataFrame único (compatibilidad v8: 1 nodo en nodo_multianos)
    #   (b) Un dict {nodo: df_multianos} para múltiples nodos
    multi_dict = None
    if isinstance(df_multianos, dict) and df_multianos:
        multi_dict = df_multianos
    elif df_multianos is not None and not df_multianos.empty and nodo_multianos:
        multi_dict = {nodo_multianos: df_multianos}

    if multi_dict:
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.marker import Marker

        meses_lbl = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

        for nodo_n, df_n in multi_dict.items():
            if df_n is None or df_n.empty:
                continue

            sheet_name = f"📊 {nodo_n[:24]}"[:31]  # Excel limit 31 chars
            ws_g = wb.create_sheet(sheet_name)

            ws_g.merge_cells("A1:N1")
            c = ws_g["A1"]
            c.value = f"PML promedio mensual por año · {nodo_n}"
            c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
            c.fill = PatternFill("solid", start_color=C_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws_g.row_dimensions[1].height = 26

            ws_g.cell(row=3, column=1, value="Mes")
            hdr(ws_g.cell(row=3, column=1, value="Mes"), bg=C_SUB)

            años_cols = sorted([c for c in df_n.columns if c != 'mes_num'])
            for ci, año in enumerate(años_cols, start=2):
                hdr(ws_g.cell(row=3, column=ci, value=str(año)), bg=C_SUB)

            for ri, mes_idx in enumerate(range(1, 13), start=4):
                ws_g.cell(row=ri, column=1, value=meses_lbl[mes_idx-1])
                ws_g.cell(row=ri, column=1).font = _FONT_DATO
                ws_g.cell(row=ri, column=1).border = BORDE
                for ci, año in enumerate(años_cols, start=2):
                    v = df_n[df_n['mes_num'] == mes_idx][año].values
                    val = float(v[0]) if len(v) > 0 and pd.notna(v[0]) else None
                    cc = ws_g.cell(row=ri, column=ci, value=val)
                    cc.font = _FONT_DATO
                    cc.border = BORDE
                    cc.alignment = _ALIGN_NUM
                    cc.number_format = "#,##0.00"

            chart = LineChart()
            chart.title = f"PML promedio mensual ({_sym}/MWh) · {nodo_n}"
            chart.style = 12
            chart.height = 12
            chart.width = 22
            chart.y_axis.title = f"PML promedio ({_sym}/MWh)"
            chart.x_axis.title = "Mes"

            data_ref = Reference(ws_g, min_col=2, max_col=1+len(años_cols),
                                  min_row=3, max_row=15)
            cats_ref = Reference(ws_g, min_col=1, min_row=4, max_row=15)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            for s in chart.series:
                s.smooth = False
                s.marker = Marker(symbol="circle", size=8)

            ws_g.add_chart(chart, "A18")

            for ci, w in enumerate([10] + [12] * len(años_cols), 1):
                ws_g.column_dimensions[get_column_letter(ci)].width = w

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# EXCEL CUSTOM — usuario selecciona qué gráficos incluir
# ═══════════════════════════════════════════════════════════════════════
def generar_excel_custom(df, df_resumen, df_metricas, opciones, sistema, proceso,
                          fecha_ini, fecha_fin, moneda="MXN", tc_info=None,
                          df_multianos_dict=None):
    """Excel personalizado con secciones que el usuario selecciona.

    opciones: dict con keys booleanas:
        'resumen', 'multiano', 'spread_ccr', 'bess_arbitraje',
        'bess_ssaa', 'bess_firming', 'top_pml', 'top_volatilidad'
    """
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.chart.marker import Marker

    _NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
    _side = Side(style="thin", color="BFBFBF")
    BORDE = Border(left=_side, right=_side, top=_side, bottom=_side)
    _sym = "USD$" if moneda == "USD" else "$"

    def hdr(cell, bg=C_HEADER, fg=C_WHITE, size=11):
        cell.font = Font(bold=True, color=fg, size=size, name="Arial")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDE

    _FONT_DATO = Font(name="Arial", size=10)
    _ALIGN_NUM = Alignment(horizontal="center", vertical="center")
    _ALIGN_TXT = Alignment(horizontal="left", vertical="center")
    _FILL_ALT  = PatternFill("solid", start_color=C_ALT)

    wb = Workbook()
    ws = wb.active
    ws.title = "■ Portada"
    ws.sheet_properties.tabColor = C_HEADER
    ws.sheet_view.showGridLines = False

    # Portada
    ws.merge_cells("B2:H4")
    c = ws["B2"]
    c.value = f"Excel Custom — Análisis PML\nGenerado a la medida"
    c.font = Font(bold=True, color=C_WHITE, size=22, name="Arial")
    c.fill = PatternFill("solid", start_color=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(2, 5):
        ws.row_dimensions[r].height = 32

    secciones_inc = [k for k, v in opciones.items() if v]
    info_rows = [
        ("Sistema", sistema), ("Proceso", proceso),
        ("Período", f"{fecha_ini} → {fecha_fin}"),
        ("Nodos", f"{df['nodo'].nunique() if not df.empty else 0}"),
        ("Moneda", f"{moneda}" + (f" · TC FIX promedio: {tc_info['tc_promedio']:.4f}"
                                     if tc_info and tc_info.get('tc_promedio') else "")),
        ("Secciones incluidas", ", ".join(secciones_inc) if secciones_inc else "—"),
        ("Fecha generación", _NOW),
    ]
    for ri, (lbl, val) in enumerate(info_rows, start=7):
        ws.merge_cells(f"B{ri}:D{ri}")
        c1 = ws[f"B{ri}"]; c1.value = lbl
        c1.font = Font(bold=True, color=C_HEADER, size=11, name="Arial")
        c1.fill = PatternFill("solid", start_color=C_INFO)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c1.border = BORDE
        ws.merge_cells(f"E{ri}:H{ri}")
        c2 = ws[f"E{ri}"]; c2.value = val
        c2.font = Font(size=11, name="Arial")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c2.border = BORDE
        ws.row_dimensions[ri].height = 22

    ws.merge_cells("B17:H17")
    c = ws["B17"]; c.value = "Prepared by: Sebastian Roldan (SRF) · Recurrent Energy"
    c.font = Font(italic=True, bold=True, color=C_HEADER, size=12, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=C_INFO)
    ws.row_dimensions[17].height = 24

    for col, w in [("A", 3), ("B", 18), ("C", 18), ("D", 18),
                   ("E", 18), ("F", 18), ("G", 18), ("H", 18), ("I", 3)]:
        ws.column_dimensions[col].width = w

    # ─── SECCIÓN: Resumen ───
    if opciones.get("resumen") and df_resumen is not None and not df_resumen.empty:
        ws_r = wb.create_sheet("📊 Resumen")
        ws_r.merge_cells("A1:K1")
        c = ws_r["A1"]; c.value = f"Resumen estadístico ({moneda})"
        c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_r.row_dimensions[1].height = 26

        cols = ["Nodo", "Nombre", "CCR", "# Reg", "Promedio", "Mediana",
                "Máximo", "Mínimo", "Volatilidad", "P95", "% horas neg"]
        for ci, col in enumerate(cols, 1):
            hdr(ws_r.cell(row=2, column=ci, value=col), bg=C_SUB)
        ws_r.row_dimensions[2].height = 22

        for i, (_, r) in enumerate(df_resumen.iterrows(), 1):
            bg = C_ALT if i % 2 == 0 else None
            valores = [r['nodo'], r['nombre'], r['ccr'], r['registros'],
                       r['promedio'], r['mediana'], r['maximo'], r['minimo'],
                       r['std'], r['p95'], r['% horas neg']]
            for ci, v in enumerate(valores, 1):
                cc = ws_r.cell(row=i+2, column=ci, value=v)
                cc.font = _FONT_DATO
                cc.border = BORDE
                cc.alignment = _ALIGN_NUM if ci > 3 else _ALIGN_TXT
                if bg: cc.fill = PatternFill("solid", start_color=bg)
                if ci in (5, 6, 7, 8, 9, 10): cc.number_format = "#,##0.00"
                elif ci == 11: cc.number_format = "0.0\"%\""

        for ci, w in enumerate([14, 22, 14, 10, 12, 12, 12, 12, 12, 12, 12], 1):
            ws_r.column_dimensions[get_column_letter(ci)].width = w
        ws_r.freeze_panes = "A3"

    # ─── SECCIÓN: BESS Scoring (3 use cases si seleccionados) ───
    bess_use_cases = []
    if opciones.get("bess_arbitraje"): bess_use_cases.append('Arbitraje')
    if opciones.get("bess_ssaa"): bess_use_cases.append('Servicios Auxiliares')
    if opciones.get("bess_firming"): bess_use_cases.append('Renewables Firming')

    for use_case in bess_use_cases:
        df_score = calcular_score_bess(df_metricas, use_case)
        if df_score.empty:
            continue
        sheet_name = f"BESS {use_case[:22]}"[:31]
        ws_b = wb.create_sheet(sheet_name)
        ws_b.merge_cells("A1:J1")
        c = ws_b["A1"]; c.value = f"BESS Scoring — {use_case}"
        c.font = Font(bold=True, color=C_WHITE, size=13, name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_b.row_dimensions[1].height = 26

        cols = ["Rank", "Nodo", "Score (0-100)", "PML promedio",
                "Volatilidad", "Spread P95-P5", "Spread día prom",
                "Cambios bruscos", "Horas pico", "% horas neg"]
        for ci, col in enumerate(cols, 1):
            hdr(ws_b.cell(row=2, column=ci, value=col), bg=C_SUB)
        ws_b.row_dimensions[2].height = 32

        for i, (_, r) in enumerate(df_score.iterrows(), 1):
            bg = C_ALT if i % 2 == 0 else None
            valores = [i, r['nodo'], r['score'], r['pml_promedio'],
                       r['volatilidad'], r['spread_p95_p5'],
                       r['spread_avg_diario'], r['cambios_bruscos'],
                       r['horas_pico'], r['pct_horas_neg']]
            for ci, v in enumerate(valores, 1):
                cc = ws_b.cell(row=i+2, column=ci, value=v)
                cc.font = _FONT_DATO
                cc.border = BORDE
                cc.alignment = _ALIGN_NUM if ci != 2 else _ALIGN_TXT
                if bg: cc.fill = PatternFill("solid", start_color=bg)
                if ci in (3, 4, 5, 6, 7): cc.number_format = "#,##0.00"
                elif ci == 10: cc.number_format = "0.0\"%\""

        for ci, w in enumerate([8, 16, 14, 14, 14, 14, 14, 14, 12, 14], 1):
            ws_b.column_dimensions[get_column_letter(ci)].width = w
        ws_b.freeze_panes = "A3"

    # ─── SECCIÓN: Multi-año (1 hoja por nodo) ───
    if opciones.get("multiano") and df_multianos_dict:
        meses_lbl = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                     'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
        for nodo_n, df_n in df_multianos_dict.items():
            if df_n is None or df_n.empty:
                continue
            sheet_name = f"📊 {nodo_n[:24]}"[:31]
            ws_g = wb.create_sheet(sheet_name)
            ws_g.merge_cells("A1:N1")
            c = ws_g["A1"]
            c.value = f"PML promedio mensual por año · {nodo_n}"
            c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
            c.fill = PatternFill("solid", start_color=C_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws_g.row_dimensions[1].height = 26

            hdr(ws_g.cell(row=3, column=1, value="Mes"), bg=C_SUB)
            años_cols = sorted([c for c in df_n.columns if c != 'mes_num'])
            for ci, año in enumerate(años_cols, start=2):
                hdr(ws_g.cell(row=3, column=ci, value=str(año)), bg=C_SUB)

            for ri, mes_idx in enumerate(range(1, 13), start=4):
                ws_g.cell(row=ri, column=1, value=meses_lbl[mes_idx-1])
                ws_g.cell(row=ri, column=1).font = _FONT_DATO
                ws_g.cell(row=ri, column=1).border = BORDE
                for ci, año in enumerate(años_cols, start=2):
                    v = df_n[df_n['mes_num'] == mes_idx][año].values
                    val = float(v[0]) if len(v) > 0 and pd.notna(v[0]) else None
                    cc = ws_g.cell(row=ri, column=ci, value=val)
                    cc.font = _FONT_DATO
                    cc.border = BORDE
                    cc.alignment = _ALIGN_NUM
                    cc.number_format = "#,##0.00"

            chart = LineChart()
            chart.title = f"PML promedio mensual ({_sym}/MWh) · {nodo_n}"
            chart.style = 12
            chart.height = 12
            chart.width = 22
            chart.y_axis.title = f"PML promedio ({_sym}/MWh)"
            chart.x_axis.title = "Mes"
            data_ref = Reference(ws_g, min_col=2, max_col=1+len(años_cols),
                                  min_row=3, max_row=15)
            cats_ref = Reference(ws_g, min_col=1, min_row=4, max_row=15)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            for s in chart.series:
                s.smooth = False
                s.marker = Marker(symbol="circle", size=8)
            ws_g.add_chart(chart, "A18")

            for ci, w in enumerate([10] + [12] * len(años_cols), 1):
                ws_g.column_dimensions[get_column_letter(ci)].width = w

    # ─── SECCIÓN: Spread por CCR ───
    if opciones.get("spread_ccr") and not df.empty:
        ccrs_unicos = sorted([c for c in df['ccr'].unique() if c and c != "?"])
        if len(ccrs_unicos) >= 2:
            ws_c = wb.create_sheet("📊 Spread CCR")
            ws_c.merge_cells("A1:N1")
            c = ws_c["A1"]
            c.value = f"Spread mensual por CCR · {len(ccrs_unicos)} zonas"
            c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
            c.fill = PatternFill("solid", start_color=C_HEADER)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws_c.row_dimensions[1].height = 26

            df_w = df[['ccr', 'fecha_dt', 'pml']].copy()
            df_w['periodo'] = df_w['fecha_dt'].dt.to_period('M').dt.to_timestamp()
            monthly = df_w.groupby(['ccr', 'periodo'])['pml'].mean().reset_index()
            pivot = monthly.pivot(index='periodo', columns='ccr', values='pml').round(2)

            hdr(ws_c.cell(row=3, column=1, value="Periodo"), bg=C_SUB)
            ccr_cols = list(pivot.columns)
            for ci, ccr in enumerate(ccr_cols, start=2):
                hdr(ws_c.cell(row=3, column=ci, value=str(ccr)), bg=C_SUB)

            for ri, (per, row_vals) in enumerate(pivot.iterrows(), start=4):
                ws_c.cell(row=ri, column=1, value=per.strftime("%b %Y"))
                ws_c.cell(row=ri, column=1).font = _FONT_DATO
                ws_c.cell(row=ri, column=1).border = BORDE
                for ci, ccr in enumerate(ccr_cols, start=2):
                    val = row_vals[ccr]
                    val_clean = float(val) if pd.notna(val) else None
                    cc = ws_c.cell(row=ri, column=ci, value=val_clean)
                    cc.font = _FONT_DATO
                    cc.border = BORDE
                    cc.alignment = _ALIGN_NUM
                    cc.number_format = "#,##0.00"

            chart_ccr = LineChart()
            chart_ccr.title = f"Spread mensual por CCR ({_sym}/MWh)"
            chart_ccr.style = 12
            chart_ccr.height = 14
            chart_ccr.width = 26
            chart_ccr.y_axis.title = f"PML promedio ({_sym}/MWh)"
            chart_ccr.x_axis.title = "Periodo"
            n_rows = len(pivot)
            data_ref = Reference(ws_c, min_col=2, max_col=1+len(ccr_cols),
                                  min_row=3, max_row=3+n_rows)
            cats_ref = Reference(ws_c, min_col=1, min_row=4, max_row=3+n_rows)
            chart_ccr.add_data(data_ref, titles_from_data=True)
            chart_ccr.set_categories(cats_ref)
            for s in chart_ccr.series:
                s.smooth = False
            ws_c.add_chart(chart_ccr, f"A{4+n_rows+2}")

            for ci, w in enumerate([14] + [14] * len(ccr_cols), 1):
                ws_c.column_dimensions[get_column_letter(ci)].width = w

    # ─── SECCIÓN: Top PML / Volatilidad ───
    if opciones.get("top_pml") or opciones.get("top_volatilidad"):
        ws_t = wb.create_sheet("🏆 Rankings")
        ws_t.merge_cells("A1:F1")
        c = ws_t["A1"]
        c.value = "Rankings comparativos"
        c.font = Font(bold=True, color=C_WHITE, size=14, name="Arial")
        c.fill = PatternFill("solid", start_color=C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_t.row_dimensions[1].height = 26

        cur_row = 3
        if opciones.get("top_pml") and not df.empty:
            ws_t.merge_cells(f"A{cur_row}:F{cur_row}")
            c = ws_t.cell(row=cur_row, column=1, value="🏆 Top PML promedio")
            c.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
            c.fill = PatternFill("solid", start_color=C_RED)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws_t.row_dimensions[cur_row].height = 22
            cur_row += 1

            top_pml = df.groupby("nodo")["pml"].mean().sort_values(ascending=False).head(10).round(2)
            for ci, h in enumerate(["Rank", "Nodo", f"PML promedio ({_sym}/MWh)"], 1):
                hdr(ws_t.cell(row=cur_row, column=ci, value=h), bg=C_SUB)
            cur_row += 1
            for i, (nodo_n, val) in enumerate(top_pml.items(), 1):
                ws_t.cell(row=cur_row, column=1, value=i).font = _FONT_DATO
                ws_t.cell(row=cur_row, column=2, value=str(nodo_n)).font = _FONT_DATO
                ws_t.cell(row=cur_row, column=3, value=val).font = _FONT_DATO
                for ci in range(1, 4):
                    ws_t.cell(row=cur_row, column=ci).border = BORDE
                    ws_t.cell(row=cur_row, column=ci).alignment = _ALIGN_NUM
                ws_t.cell(row=cur_row, column=3).number_format = "#,##0.00"
                cur_row += 1
            cur_row += 1

        if opciones.get("top_volatilidad") and not df.empty:
            ws_t.merge_cells(f"A{cur_row}:F{cur_row}")
            c = ws_t.cell(row=cur_row, column=1, value="📊 Top volatilidad")
            c.font = Font(bold=True, color=C_WHITE, size=12, name="Arial")
            c.fill = PatternFill("solid", start_color=C_RED)
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws_t.row_dimensions[cur_row].height = 22
            cur_row += 1

            top_vol = df.groupby("nodo")["pml"].std().sort_values(ascending=False).head(10).round(2)
            for ci, h in enumerate(["Rank", "Nodo", f"Volatilidad ({_sym}/MWh)"], 1):
                hdr(ws_t.cell(row=cur_row, column=ci, value=h), bg=C_SUB)
            cur_row += 1
            for i, (nodo_n, val) in enumerate(top_vol.items(), 1):
                ws_t.cell(row=cur_row, column=1, value=i).font = _FONT_DATO
                ws_t.cell(row=cur_row, column=2, value=str(nodo_n)).font = _FONT_DATO
                ws_t.cell(row=cur_row, column=3, value=val).font = _FONT_DATO
                for ci in range(1, 4):
                    ws_t.cell(row=cur_row, column=ci).border = BORDE
                    ws_t.cell(row=cur_row, column=ci).alignment = _ALIGN_NUM
                ws_t.cell(row=cur_row, column=3).number_format = "#,##0.00"
                cur_row += 1

        for ci, w in enumerate([8, 18, 22, 14, 14, 14], 1):
            ws_t.column_dimensions[get_column_letter(ci)].width = w

    buffer = io.BytesIO()
    wb.save(buffer); buffer.seek(0)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# KMZ
# ═══════════════════════════════════════════════════════════════════════
def generar_kmz(matches_df):
    KML_NS = 'http://www.opengis.net/kml/2.2'
    def _se(parent, tag, text=None, **attrib):
        el = ET.SubElement(parent, tag, attrib)
        if text is not None: el.text = str(text)
        return el

    root = ET.Element('kml', xmlns=KML_NS)
    doc = _se(root, 'Document')
    _se(doc, 'name', f'PML CENACE Geo — {datetime.now().strftime("%Y-%m-%d")} · SRF')
    _se(doc, 'open', '1')

    def _style(parent, sid, color_abgr, scale=1.2):
        st_el = _se(parent, 'Style', id=sid)
        ic = _se(st_el, 'IconStyle')
        _se(ic, 'color', color_abgr); _se(ic, 'scale', str(scale))
        ico = _se(ic, 'Icon')
        _se(ico, 'href', 'http://maps.google.com/mapfiles/kml/shapes/square.png')
        lbl = _se(st_el, 'LabelStyle'); _se(lbl, 'scale', '0.85')
        return st_el

    _style(doc, 'styleExc',     'ff00aa00', 1.3)
    _style(doc, 'styleBueno',   'ff00bbff', 1.2)
    _style(doc, 'styleAcept',   'ff0088ff', 1.1)

    folder_main = _se(doc, 'Folder')
    _se(folder_main, 'name', '⚡ Nodos CENACE')
    _se(folder_main, 'open', '1')

    matches_df_ok = matches_df[matches_df['lat'].notna()].copy()
    regiones_count = {}
    for _, row in matches_df_ok.iterrows():
        ccr = str(row['ccr']) or 'SIN CLASIFICAR'
        regiones_count[ccr] = regiones_count.get(ccr, 0) + 1

    folders = {}
    for ccr in sorted(regiones_count, key=lambda r: -regiones_count[r]):
        f = _se(folder_main, 'Folder')
        _se(f, 'name', f'📍 {ccr} ({regiones_count[ccr]})')
        _se(f, 'open', '0')
        folders[ccr] = {}
        for cal in ['🥇 Excelente', '🥈 Bueno', '🥉 Aceptable']:
            sf = _se(f, 'Folder')
            _se(sf, 'name', cal)
            _se(sf, 'open', '1' if 'Excelente' in cal else '0')
            folders[ccr][cal] = sf

    for _, row in matches_df_ok.iterrows():
        cal = str(row['calidad'])
        ccr = str(row['ccr']) or 'SIN CLASIFICAR'
        folder = folders.get(ccr, {}).get(cal, folder_main)
        if cal.startswith('🥇'):    style = '#styleExc'
        elif cal.startswith('🥈'):  style = '#styleBueno'
        else:                      style = '#styleAcept'

        pm = _se(folder, 'Placemark')
        _se(pm, 'name', f"{row['clave']} · {str(row['nombre_osm'])[:25]}")
        _se(pm, 'styleUrl', style)
        desc = (
            f"<b>Clave CENACE:</b> {row['clave']}<br/>"
            f"<b>Nombre CENACE:</b> {row['nombre_cenace']}<br/>"
            f"<b>CCR / Zona:</b> {row['ccr']} · {row['zona']}<br/>"
            f"<b>Voltaje CENACE:</b> {row['kv_cenace']} kV<br/>"
            f"<b>Estado / Municipio:</b> {row['estado']} · {row['municipio']}<br/><br/>"
            f"<b>━━━ Match OSM ━━━</b><br/>"
            f"<b>Calidad:</b> {row['calidad']} (sim {row['similitud']})<br/>"
            f"<b>Nombre OSM:</b> {row['nombre_osm']}<br/>"
            f"<b>Voltaje OSM:</b> {row['kv_osm']} kV<br/>"
            f"<b>Operador:</b> {row['operator_osm']}<br/>"
            f"<i>SRF · Recurrent Energy</i>"
        )
        _se(pm, 'description', desc)
        pt = _se(pm, 'Point')
        _se(pt, 'coordinates', f"{float(row['lon'])},{float(row['lat'])},0")

    kml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + ET.tostring(root, encoding='unicode')
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('doc.kml', kml_str.encode('utf-8'))
    buffer.seek(0)
    return buffer.getvalue()


# ═══════════════════════════════════════════════════════════════════════
# DATAFRAMES (cached para performance)
# ═══════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False, max_entries=3)
def acumulado_a_dataframe_cached(acumulado_id, _acumulado, _catalogo):
    """Convierte acumulado a DF con tipos optimizados para memoria."""
    rows = []
    for nodo, filas in _acumulado.items():
        info = _catalogo.get(nodo, {}) if _catalogo else {}
        ccr = info.get("ccr", "?")
        nombre = info.get("nombre", nodo)
        estado = info.get("estado", "")
        for f in filas:
            try: pml_val = float(f["pml"])
            except (TypeError, ValueError): continue
            rows.append({
                "nodo":   nodo,
                "ccr":    ccr,
                "nombre": nombre,
                "estado": estado,
                "fecha":  f["fecha"],
                "hora":   int(f["hora"]) if f["hora"] else 0,
                "pml":    pml_val,
            })
    df = pd.DataFrame(rows)
    rows = None  # liberar memoria de la lista
    gc.collect()

    if not df.empty:
        # Convertir a categorical para ahorrar memoria (string columnas repetitivas)
        df["nodo"] = df["nodo"].astype('category')
        df["ccr"] = df["ccr"].astype('category')
        df["nombre"] = df["nombre"].astype('category')
        df["estado"] = df["estado"].astype('category')
        df["hora"] = df["hora"].astype('int8')
        df["pml"] = df["pml"].astype('float32')  # vs float64 por defecto = 50% menos memoria

        df["fecha_dt"] = pd.to_datetime(df["fecha"], errors="coerce")
        df["mes"] = df["fecha_dt"].dt.month.astype('int8')
        df["mes_nombre"] = df["fecha_dt"].dt.strftime("%b %Y").astype('category')

        # Drop la columna fecha string (redundante con fecha_dt) — ahorra mucha memoria
        df = df.drop(columns=["fecha"])
        gc.collect()
    return df


def acumulado_a_dataframe(acumulado, catalogo):
    """Wrapper con hash key para cache."""
    if not acumulado: return pd.DataFrame()
    # Hash key: # nodos + total filas + primer nodo
    key = (len(acumulado),
           sum(len(v) for v in acumulado.values()),
           list(acumulado.keys())[0] if acumulado else "")
    return acumulado_a_dataframe_cached(str(key), acumulado, catalogo)


@st.cache_data(show_spinner=False, max_entries=3)
def calcular_resumen_cached(df_hash, _df):
    if _df.empty: return pd.DataFrame()
    summary = _df.groupby("nodo").agg(
        nombre=("nombre", "first"),
        ccr=("ccr", "first"),
        registros=("pml", "count"),
        promedio=("pml", "mean"),
        mediana=("pml", "median"),
        maximo=("pml", "max"),
        minimo=("pml", "min"),
        std=("pml", "std"),
    ).round(2)
    pct_neg = _df.groupby("nodo")["pml"].apply(
        lambda x: (x < 0).sum() / len(x) * 100 if len(x) > 0 else 0
    ).round(1)
    summary["% horas neg"] = pct_neg
    summary["p95"] = _df.groupby("nodo")["pml"].quantile(0.95).round(2)
    summary["p5"] = _df.groupby("nodo")["pml"].quantile(0.05).round(2)
    return summary.reset_index().sort_values("promedio", ascending=False).reset_index(drop=True)


def calcular_resumen(df):
    if df.empty: return pd.DataFrame()
    key = f"{len(df)}-{df['pml'].sum():.0f}"
    return calcular_resumen_cached(key, df)


# Layout estándar para gráficas
def _layout_estandar(titulo, height=480):
    return dict(
        title=dict(
            text=titulo,
            font=dict(family="Arial Black", size=16, color=TEXT_TITLE),
            x=0.02, xanchor="left",
        ),
        height=height,
        plot_bgcolor="white",
        paper_bgcolor="white",
        font=dict(family="Arial", size=12, color=TEXT_DARK),
        margin=dict(l=70, r=40, t=70, b=60),
        hoverlabel=dict(bgcolor="white", font_size=12, font_color=TEXT_DARK,
                        bordercolor=AXIS_LINE),
        uirevision='constant',  # mantiene zoom/pan al actualizar
    )


def _ejes_estandar(fig, x_title=None, y_title=None):
    fig.update_xaxes(
        showgrid=True, gridcolor=GRID_LIGHT,
        linecolor=AXIS_LINE, linewidth=1.5,
        tickfont=dict(family="Arial", size=11, color=TEXT_DARK),
        title=dict(text=x_title or "", font=dict(family="Arial", size=13, color=TEXT_DARK)),
    )
    fig.update_yaxes(
        showgrid=True, gridcolor=GRID_LIGHT,
        linecolor=AXIS_LINE, linewidth=1.5,
        tickfont=dict(family="Arial", size=11, color=TEXT_DARK),
        title=dict(text=y_title or "", font=dict(family="Arial", size=13, color=TEXT_DARK)),
        zeroline=True, zerolinecolor=RE_RED, zerolinewidth=1.5,
    )
    return fig


def grafica_lineas_tiempo(df, max_nodos=15):
    if df.empty: return None
    promedios = df.groupby("nodo")["pml"].mean().sort_values(ascending=False)
    nodos_mostrar = promedios.head(max_nodos).index.tolist()
    df_plot = df[df["nodo"].isin(nodos_mostrar)].copy()
    daily = df_plot.groupby(["nodo", "fecha_dt"])["pml"].mean().reset_index()
    fig = px.line(
        daily, x="fecha_dt", y="pml", color="nodo",
        color_discrete_sequence=PALETTE,
    )
    fig.update_traces(line=dict(width=2.5), mode='lines')
    fig.update_layout(**_layout_estandar(
        f"Evolución temporal · Top {len(nodos_mostrar)} nodos por PML promedio"))
    fig.update_layout(
        hovermode="x unified",
        legend=dict(
            orientation="v", yanchor="top", y=1, xanchor="left", x=1.02,
            bgcolor="rgba(255,255,255,0.95)",
            bordercolor=AXIS_LINE, borderwidth=1,
            font=dict(family="Arial", size=11, color=TEXT_DARK),
            title=dict(text="<b>Nodo</b>", font=dict(color=TEXT_DARK)),
        ),
        margin=dict(l=70, r=200, t=70, b=60),
    )
    return _ejes_estandar(fig, "Fecha", f"PML promedio diario {label_moneda()}")


def grafica_heatmap_horario(df, nodo_seleccionado):
    if df.empty: return None
    sub = df[df["nodo"] == nodo_seleccionado].copy()
    if sub.empty: return None
    pivot = sub.pivot_table(values="pml", index="hora", columns="mes_nombre",
                              aggfunc="mean").round(1)
    if not pivot.empty:
        meses_orden = sub.sort_values("fecha_dt")["mes_nombre"].drop_duplicates().tolist()
        pivot = pivot.reindex(columns=meses_orden)

    colorscale = [
        [0.0, "#1a5490"], [0.25, "#5ba0d4"], [0.5, "#FFFFFF"],
        [0.75, "#f5a653"], [1.0, "#a0090c"],
    ]
    fig = go.Figure(data=go.Heatmap(
        z=pivot.values, x=pivot.columns, y=pivot.index,
        colorscale=colorscale,
        zmid=pivot.values.mean() if pivot.size > 0 else 0,
        colorbar=dict(
            title=dict(text=f"{label_moneda_short()}/MWh",
                       font=dict(color=TEXT_DARK, size=12)),
            thickness=15, len=0.85,
            tickfont=dict(color=TEXT_DARK, size=11),
        ),
        hovertemplate=(f"<b>Hora:</b> %{{y}}h<br><b>Mes:</b> %{{x}}<br>"
                       f"<b>PML:</b> {label_moneda_short()}%{{z}}<extra></extra>"),
        text=pivot.values,
        texttemplate="%{text:.0f}",
        textfont=dict(size=10, color=TEXT_DARK, family="Arial Black"),
    ))
    fig.update_layout(**_layout_estandar(
        f"Heatmap PML hora × mes · {nodo_seleccionado}", height=540))
    fig.update_yaxes(autorange="reversed", dtick=2,
                      linecolor=AXIS_LINE, linewidth=1.5,
                      tickfont=dict(family="Arial", size=11, color=TEXT_DARK),
                      title=dict(text="Hora del día", font=dict(family="Arial", size=13, color=TEXT_DARK)))
    fig.update_xaxes(linecolor=AXIS_LINE, linewidth=1.5,
                      tickfont=dict(family="Arial", size=11, color=TEXT_DARK),
                      title=dict(text="Mes", font=dict(family="Arial", size=13, color=TEXT_DARK)))
    return fig


def grafica_pml_multiano(df, nodo_seleccionado):
    """Gráfica de PML promedio mensual por año (estilo SRF screenshot).

    Para el nodo seleccionado, calcula el PML promedio de cada mes en cada año
    presente en los datos. Genera puntos conectados con colores corporativos.
    """
    if df.empty: return None, []
    sub = df[df["nodo"] == nodo_seleccionado].copy()
    if sub.empty: return None, []

    sub["año"] = sub["fecha_dt"].dt.year
    sub["mes_num"] = sub["fecha_dt"].dt.month

    años_disponibles = sorted(sub["año"].unique())
    if len(años_disponibles) < 1:
        return None, []

    # Promedio mensual por año
    pivot = sub.pivot_table(values="pml", index="mes_num", columns="año",
                              aggfunc="mean").round(1)

    # Reindexar para tener todos los 12 meses
    pivot = pivot.reindex(range(1, 13))

    meses_labels = ['ene', 'feb', 'mar', 'abr', 'may', 'jun',
                    'jul', 'ago', 'sep', 'oct', 'nov', 'dic']

    # Colores corporativos por año (rotando paleta)
    color_anos = {
        años_disponibles[0]: RE_RED,
    }
    if len(años_disponibles) >= 2:
        color_anos[años_disponibles[1]] = "#5ba0d4"  # azul claro RE
    if len(años_disponibles) >= 3:
        color_anos[años_disponibles[2]] = RE_NAVY
    if len(años_disponibles) >= 4:
        color_anos[años_disponibles[3]] = "#d4a017"
    # Más años: rotar paleta
    for i, year in enumerate(años_disponibles[4:]):
        color_anos[year] = PALETTE[(i + 4) % len(PALETTE)]

    fig = go.Figure()
    promedios_anuales = {}

    for año in años_disponibles:
        if año not in pivot.columns:
            continue
        valores = pivot[año]
        promedio_anual = valores.mean()
        promedios_anuales[año] = promedio_anual

        sym = label_moneda_short()
        nombre_traza = f"PML Avg {año}: {sym}{promedio_anual:.0f}/MWh"

        fig.add_trace(go.Scatter(
            x=meses_labels,
            y=valores.values,
            mode='lines+markers',
            name=nombre_traza,
            line=dict(color=color_anos[año], width=2, dash='dot'),
            marker=dict(color=color_anos[año], size=10,
                        line=dict(color='white', width=1.5)),
            hovertemplate=f"<b>{año}</b><br>%{{x}}: {sym}%{{y:.1f}}<extra></extra>",
        ))

    fig.update_layout(**_layout_estandar(
        f"📊 PML promedio mensual · {nodo_seleccionado}", height=520))

    fig.update_layout(
        legend=dict(
            orientation="v", yanchor="top", y=0.98, xanchor="right", x=0.98,
            bgcolor="rgba(255,255,255,0.95)",
            bordercolor=AXIS_LINE, borderwidth=1.5,
            font=dict(family="Arial", size=12, color=TEXT_DARK),
            itemsizing='constant',  # mantiene el dot del color visible
        ),
        margin=dict(l=70, r=40, t=70, b=60),
    )
    return _ejes_estandar(fig, "Mes", f"PML promedio {label_moneda()}"), años_disponibles


def grafica_spread_ccr(df, granularidad="mensual"):
    """Gráfica de spread por CCR — una línea continua por zona en el tiempo.

    Útil cuando hay nodos de varios CCRs en la consulta.
    Promedia todos los nodos de cada CCR por mes.

    Retorna (fig, ccrs_disponibles).
    """
    if df.empty:
        return None, []

    ccrs_unicos = sorted([c for c in df['ccr'].unique() if c and c != "?"])
    if len(ccrs_unicos) < 2:
        return None, ccrs_unicos

    # Pre-agregar a mensual para reducir memoria y mejorar legibilidad
    df_work = df[['ccr', 'fecha_dt', 'pml']].copy()
    df_work['periodo'] = df_work['fecha_dt'].dt.to_period('M').dt.to_timestamp()
    monthly = df_work.groupby(['ccr', 'periodo'])['pml'].mean().reset_index()

    # Asignar colores corporativos por CCR (consistente con paleta)
    color_ccr = {ccr: PALETTE[i % len(PALETTE)] for i, ccr in enumerate(ccrs_unicos)}

    fig = go.Figure()
    sym = label_moneda_short()

    for ccr in ccrs_unicos:
        sub = monthly[monthly['ccr'] == ccr].sort_values('periodo')
        if sub.empty:
            continue
        promedio = sub['pml'].mean()
        fig.add_trace(go.Scatter(
            x=sub['periodo'],
            y=sub['pml'],
            mode='lines+markers',
            name=f"{ccr} (avg {sym}{promedio:.0f})",
            line=dict(color=color_ccr[ccr], width=2.5),
            marker=dict(color=color_ccr[ccr], size=7,
                        line=dict(color='white', width=1)),
            hovertemplate=f"<b>{ccr}</b><br>%{{x|%b %Y}}: {sym}%{{y:.1f}}<extra></extra>",
        ))

    fig.update_layout(**_layout_estandar(
        f"📊 Spread mensual por CCR · {len(ccrs_unicos)} zonas comparadas", height=520))
    fig.update_layout(
        legend=dict(
            orientation="v", yanchor="top", y=0.98, xanchor="right", x=0.98,
            bgcolor="rgba(255,255,255,0.95)",
            bordercolor=AXIS_LINE, borderwidth=1.5,
            font=dict(family="Arial", size=11, color=TEXT_DARK),
        ),
        hovermode='x unified',
    )
    return _ejes_estandar(fig, "Mes",
                          f"PML promedio mensual {label_moneda()}"), ccrs_unicos


def grafica_barras_top(df, metrica="promedio", top_n=10):
    if df.empty: return None
    summary = df.groupby("nodo")["pml"].agg(["mean", "std", "max", "min"])
    summary["pct_neg"] = df.groupby("nodo")["pml"].apply(lambda x: (x < 0).sum() / len(x) * 100)
    summary = summary.reset_index()
    if metrica == "promedio":
        col, titulo, color_bar = "mean", f"Top {top_n} · Mayor PML promedio", RE_NAVY
    elif metrica == "volatilidad":
        col, titulo, color_bar = "std", f"Top {top_n} · Mayor volatilidad", RE_RED
    elif metrica == "negativos":
        col, titulo, color_bar = "pct_neg", f"Top {top_n} · Más horas negativas", RE_BLUE
    else: return None
    top = summary.nlargest(top_n, col).round(2)

    text_template = "%{x:.2f}" if metrica != "negativos" else "%{x:.1f}%"
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=top[col], y=top["nodo"],
        orientation='h',
        marker=dict(color=color_bar, line=dict(color=TEXT_DARK, width=1)),
        text=top[col],
        texttemplate=text_template,
        textposition='outside',
        textfont=dict(family="Arial Black", size=11, color=TEXT_DARK),
        hovertemplate=f"<b>%{{y}}</b><br>{metrica.capitalize()}: %{{x:.2f}}<extra></extra>",
    ))
    fig.update_layout(**_layout_estandar(titulo, height=440))
    fig.update_layout(yaxis=dict(autorange="reversed"))
    x_title = metrica.capitalize() + f" {label_moneda()}" if metrica != "negativos" else "% de horas"
    return _ejes_estandar(fig, x_title, "")


# ═══════════════════════════════════════════════════════════════════════
# BESS SCORING
# ═══════════════════════════════════════════════════════════════════════
PESOS_BESS = {
    'Arbitraje': {
        'spread_p95_p5':    0.30,
        'spread_avg_diario': 0.25,
        'volatilidad':      0.20,
        'horas_pico':       0.15,
        'pml_promedio':     0.10,
    },
    'Servicios Auxiliares': {
        'volatilidad':      0.40,
        'cambios_bruscos':  0.30,
        'pml_promedio':     0.15,
        'spread_p95_p5':    0.15,
    },
    'Renewables Firming': {
        'pct_horas_neg':    0.35,
        'spread_dia':       0.30,
        'spread_p95_p5':    0.20,
        'pml_promedio':     0.15,
    },
}

DESCRIPCIONES_USE_CASE = {
    'Arbitraje': '💰 **Arbitraje energético** — Comprar barato, vender caro. '
                  'Premia spread alto, volatilidad y picos de precio.',
    'Servicios Auxiliares': '⚙️ **Servicios Auxiliares (SSAA)** — Regulación de '
                             'frecuencia y voltaje. Premia volatilidad y cambios '
                             'bruscos en el precio.',
    'Renewables Firming': '🌅 **Firming de Renovables** — Acompañar generación solar/eólica. '
                           'Premia horas con precios negativos y diferencial día-noche.',
}

DESCRIPCIONES_USE_CASE_PLAIN = {
    'Arbitraje': 'Arbitraje energético — Premia spread alto, volatilidad y picos de precio.',
    'Servicios Auxiliares': 'Servicios Auxiliares — Premia volatilidad y cambios bruscos.',
    'Renewables Firming': 'Renewables Firming — Premia horas negativas y diferencial día-noche.',
}


@st.cache_data(show_spinner=False, max_entries=3)
def calcular_metricas_bess_cached(df_hash, _df):
    if _df.empty: return pd.DataFrame()
    metricas = []
    for nodo in _df["nodo"].unique():
        sub = _df[_df["nodo"] == nodo].copy()
        pml = sub["pml"].values
        if len(pml) == 0: continue

        prom = pml.mean()
        std_v = pml.std()
        p95 = np.percentile(pml, 95)
        p5  = np.percentile(pml, 5)
        spread_p95_p5 = p95 - p5

        sub["fecha_only"] = sub["fecha_dt"].dt.date
        daily = sub.groupby("fecha_only")["pml"].agg(['min', 'max', 'mean'])
        daily["spread"] = daily["max"] - daily["min"]
        spread_avg_diario = daily["spread"].mean() if not daily.empty else 0
        spread_dia = daily["max"].mean() - daily["min"].mean() if not daily.empty else 0

        sub_sorted = sub.sort_values(["fecha_dt", "hora"])
        diffs = sub_sorted["pml"].diff().abs()
        threshold = std_v * 1.5 if std_v > 0 else 0
        cambios_bruscos = (diffs > threshold).sum()

        threshold_pico = p95
        horas_pico = (pml > threshold_pico).sum()

        pct_neg = (pml < 0).mean() * 100

        metricas.append({
            'nodo':          nodo,
            'pml_promedio':  round(prom, 2),
            'volatilidad':   round(std_v, 2),
            'spread_p95_p5': round(spread_p95_p5, 2),
            'spread_avg_diario': round(spread_avg_diario, 2),
            'spread_dia':    round(spread_dia, 2),
            'cambios_bruscos': int(cambios_bruscos),
            'horas_pico':    int(horas_pico),
            'pct_horas_neg': round(pct_neg, 1),
        })

    return pd.DataFrame(metricas)


def calcular_metricas_bess(df):
    if df.empty: return pd.DataFrame()
    key = f"{len(df)}-{df['pml'].sum():.0f}"
    return calcular_metricas_bess_cached(key, df)


def calcular_score_bess(df_metricas, use_case, pesos=None):
    """Score con rank-percentile."""
    if df_metricas.empty: return pd.DataFrame()
    if pesos is None: pesos = PESOS_BESS.get(use_case, PESOS_BESS['Arbitraje'])

    df = df_metricas.copy()
    score = pd.Series(0.0, index=df.index)
    for metrica, peso in pesos.items():
        if metrica not in df.columns: continue
        ranks = df[metrica].rank(pct=True, method='average')
        score += ranks * peso

    df['score'] = (score * 100).round(1)
    return df.sort_values('score', ascending=False).reset_index(drop=True)


def grafica_bess_ranking(df_score, use_case, top_n=15):
    if df_score.empty: return None
    top = df_score.nlargest(top_n, 'score').copy()

    colors = []
    for s in top['score']:
        if s >= 75: colors.append("#1a8a3a")
        elif s >= 50: colors.append("#9bc24f")
        elif s >= 25: colors.append("#f5d017")
        else: colors.append("#a0090c")

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=top['score'], y=top['nodo'],
        orientation='h',
        marker=dict(color=colors, line=dict(color=TEXT_DARK, width=1)),
        text=top['score'],
        texttemplate="%{x:.1f}",
        textposition='outside',
        textfont=dict(family="Arial Black", size=12, color=TEXT_DARK),
        hovertemplate="<b>%{y}</b><br>Score BESS: <b>%{x:.1f}</b>/100<extra></extra>",
    ))
    fig.update_layout(**_layout_estandar(
        f"🏆 Top {top_n} nodos · Score BESS para {use_case}", height=520))
    fig.update_layout(yaxis=dict(autorange="reversed"))
    return _ejes_estandar(fig, "Score BESS (0-100)", "")


# ═══════════════════════════════════════════════════════════════════════
# MAPA INTERACTIVO con LEYENDA CCR
# ═══════════════════════════════════════════════════════════════════════
def _ccr_color_map(matches_df):
    """Asigna un color de PALETTE a cada CCR único en el dataframe."""
    ccrs_unique = sorted(matches_df['ccr'].fillna('?').unique())
    return {ccr: PALETTE[i % len(PALETTE)] for i, ccr in enumerate(ccrs_unique)}


def grafica_mapa(matches_df, df_pml=None, color_by='pml'):
    """color_by: 'pml' (con datos) o 'ccr' (modo solo mapa)."""
    matches_ok = matches_df[matches_df['lat'].notna()].copy()
    if matches_ok.empty:
        return None

    if color_by == 'pml' and df_pml is not None and not df_pml.empty:
        pml_avg = df_pml.groupby('nodo')['pml'].agg(['mean', 'count', 'max', 'min']).reset_index()
        pml_avg.columns = ['clave', 'pml_avg', 'pml_count', 'pml_max', 'pml_min']
        map_df = matches_ok.merge(pml_avg, on='clave', how='left')

        map_df['hover'] = map_df.apply(
            lambda r: (
                f"<b>{r['clave']}</b><br>"
                f"<b>Nombre:</b> {r['nombre_cenace']}<br>"
                f"<b>CCR:</b> {r['ccr']} · {r['zona']}<br>"
                f"<b>Voltaje:</b> {r['kv_cenace']} kV<br>"
                f"<b>Estado:</b> {r['estado']}<br>"
                f"<b>Municipio:</b> {r['municipio']}<br>"
                f"━━━━━━━━━━━<br>"
                f"<b>PML promedio:</b> {label_moneda_short()}{r['pml_avg']:.2f}<br>"
                f"<b>PML máx:</b> {label_moneda_short()}{r['pml_max']:.2f}<br>"
                f"<b>PML mín:</b> {label_moneda_short()}{r['pml_min']:.2f}<br>"
                f"━━━━━━━━━━━<br>"
                f"<b>Match:</b> {r['calidad']} (sim {r['similitud']})<br>"
                f"<b>OSM:</b> {r['nombre_osm']}"
            ), axis=1
        )
        marker_color = map_df['pml_avg']
        cmin = map_df['pml_avg'].min()
        cmax = map_df['pml_avg'].max()
        colorscale = [
            [0.0, "#1a8a3a"], [0.25, "#9bc24f"], [0.5, "#f5d017"],
            [0.75, "#f57f17"], [1.0, "#a0090c"],
        ]
        colorbar = dict(
            title=dict(text=f"<b>PML promedio<br>{label_moneda()}</b>",
                       font=dict(family="Arial", size=12, color=TEXT_DARK)),
            thickness=18, len=0.7, x=1.02,
            tickfont=dict(family="Arial", size=11, color=TEXT_DARK),
        )
        showscale = True
    else:
        # Modo solo mapa: colorear por CCR
        map_df = matches_ok.copy()
        ccr_to_color = _ccr_color_map(map_df)
        map_df['hover'] = map_df.apply(
            lambda r: (
                f"<b>{r['clave']}</b><br>"
                f"<b>Nombre:</b> {r['nombre_cenace']}<br>"
                f"<b>CCR:</b> {r['ccr']} · {r['zona']}<br>"
                f"<b>Voltaje:</b> {r['kv_cenace']} kV<br>"
                f"<b>Estado:</b> {r['estado']}<br>"
                f"<b>Municipio:</b> {r['municipio']}<br>"
                f"━━━━━━━━━━━<br>"
                f"<b>Match:</b> {r['calidad']} (sim {r['similitud']})<br>"
                f"<b>OSM:</b> {r['nombre_osm']}<br>"
                f"<i>(modo solo mapa — sin datos PML)</i>"
            ), axis=1
        )
        map_df['color_ccr'] = map_df['ccr'].fillna('?').map(ccr_to_color)
        marker_color = map_df['color_ccr']
        cmin = cmax = None
        colorscale = None
        colorbar = None
        showscale = False

    # Auto-zoom
    lat_min, lat_max = map_df['lat'].min(), map_df['lat'].max()
    lon_min, lon_max = map_df['lon'].min(), map_df['lon'].max()
    lat_center = (lat_min + lat_max) / 2
    lon_center = (lon_min + lon_max) / 2
    spread = max(lat_max - lat_min, lon_max - lon_min)
    if spread > 15: zoom = 4
    elif spread > 8: zoom = 5
    elif spread > 4: zoom = 6
    elif spread > 2: zoom = 7
    elif spread > 1: zoom = 8
    else: zoom = 9

    fig = go.Figure()

    # Capa 1: halo blanco grande
    fig.add_trace(go.Scattermapbox(
        lat=map_df['lat'], lon=map_df['lon'],
        mode='markers',
        marker=dict(size=28, color='white', opacity=0.95),
        hoverinfo='skip', showlegend=False,
    ))

    # Capa 2: marker con color
    if color_by == 'pml' and showscale:
        fig.add_trace(go.Scattermapbox(
            lat=map_df['lat'], lon=map_df['lon'],
            mode='markers+text',
            marker=dict(
                size=20, color=marker_color,
                colorscale=colorscale, cmin=cmin, cmax=cmax,
                colorbar=colorbar, showscale=True,
                opacity=0.95,
            ),
            text=map_df['clave'], textposition="top right",
            textfont=dict(size=11, color=TEXT_TITLE, family="Arial Black"),
            hovertext=map_df['hover'], hoverinfo='text',
            name='Nodos', showlegend=False,
        ))
    else:
        fig.add_trace(go.Scattermapbox(
            lat=map_df['lat'], lon=map_df['lon'],
            mode='markers+text',
            marker=dict(size=20, color=marker_color, opacity=0.95),
            text=map_df['clave'], textposition="top right",
            textfont=dict(size=11, color=TEXT_TITLE, family="Arial Black"),
            hovertext=map_df['hover'], hoverinfo='text',
            name='Nodos', showlegend=False,
        ))

    # Capa 3: punto central pequeño negro
    fig.add_trace(go.Scattermapbox(
        lat=map_df['lat'], lon=map_df['lon'],
        mode='markers',
        marker=dict(size=4, color=TEXT_DARK, opacity=1.0),
        hoverinfo='skip', showlegend=False,
    ))

    titulo_mapa = (f"📍 Mapa interactivo · {len(map_df)} nodos · "
                   f"{'colores por PML' if color_by == 'pml' else 'colores por CCR'}")

    fig.update_layout(
        mapbox=dict(
            style='open-street-map',
            center=dict(lat=lat_center, lon=lon_center),
            zoom=zoom,
        ),
        height=680,
        margin=dict(l=0, r=0, t=50, b=0),
        title=dict(
            text=titulo_mapa,
            font=dict(family="Arial Black", size=16, color=TEXT_TITLE),
            x=0.02, xanchor="left",
        ),
        showlegend=False,
        paper_bgcolor="white",
        hoverlabel=dict(bgcolor="white", font_size=12, font_color=TEXT_DARK,
                        bordercolor=AXIS_LINE),
        uirevision='constant',
    )
    return fig


def render_leyenda_ccr(matches_df):
    """Leyenda de CCR debajo del mapa con colores y conteo."""
    matches_ok = matches_df[matches_df['lat'].notna()].copy()
    if matches_ok.empty: return

    ccr_to_color = _ccr_color_map(matches_ok)
    counts = matches_ok['ccr'].fillna('?').value_counts()

    items_html = ""
    for ccr in sorted(ccr_to_color.keys()):
        color = ccr_to_color[ccr]
        count = counts.get(ccr, 0)
        items_html += (
            f'<span class="ccr-legend-item">'
            f'<span class="ccr-dot" style="background:{color}"></span>'
            f'<b>{ccr}</b> ({count})'
            f'</span>'
        )

    legend_html = (
        f'<div class="ccr-legend">'
        f'<b style="color:{TEXT_DARK};font-size:0.92rem;">📍 CCRs en esta consulta:</b><br>'
        f'{items_html}'
        f'</div>'
    )
    st.markdown(legend_html, unsafe_allow_html=True)


def render_panel_ccr(matches_df, df_pml=None):
    matches_ok = matches_df[matches_df['lat'].notna()].copy()
    if matches_ok.empty: return

    if df_pml is not None and not df_pml.empty:
        pml_avg_nodo = df_pml.groupby('nodo')['pml'].mean().reset_index()
        pml_avg_nodo.columns = ['clave', 'pml_avg']
        merged = matches_ok.merge(pml_avg_nodo, on='clave', how='left')

        ccr_stats = merged.groupby('ccr').agg(
            nodos=('clave', 'count'),
            pml_promedio=('pml_avg', 'mean'),
            pml_max=('pml_avg', 'max'),
            pml_min=('pml_avg', 'min'),
        ).round(2).reset_index().sort_values('pml_promedio', ascending=False)

        st.markdown("##### 📊 Promedio por CCR (zonas analizadas)")
        st.dataframe(
            ccr_stats, use_container_width=True, hide_index=True,
            column_config={
                "ccr":          st.column_config.TextColumn("CCR", width="medium"),
                "nodos":        st.column_config.NumberColumn("# Nodos", format="%d"),
                "pml_promedio": st.column_config.NumberColumn(f"PML promedio", format=fmt_moneda()),
                "pml_max":      st.column_config.NumberColumn("Nodo más caro", format=fmt_moneda()),
                "pml_min":      st.column_config.NumberColumn("Nodo más barato", format=fmt_moneda()),
            },
        )
    else:
        ccr_stats = matches_ok.groupby('ccr').agg(
            nodos=('clave', 'count'),
        ).reset_index().sort_values('nodos', ascending=False)

        st.markdown("##### 📊 Conteo por CCR")
        st.dataframe(
            ccr_stats, use_container_width=True, hide_index=True,
            column_config={
                "ccr":   st.column_config.TextColumn("CCR", width="medium"),
                "nodos": st.column_config.NumberColumn("# Nodos", format="%d"),
            },
        )


# ═══════════════════════════════════════════════════════════════════════
# RENDER BESS SCORING
# ═══════════════════════════════════════════════════════════════════════
def render_bess_scoring(df, use_case_default='Arbitraje'):
    st.divider()
    st.markdown("## 🔋 BESS Scoring")
    st.caption("Ranking de nodos por idoneidad para distintos casos de uso. "
               "Score 0-100 usando rank-percentile.")

    use_case = st.selectbox(
        "Caso de uso",
        list(PESOS_BESS.keys()),
        index=list(PESOS_BESS.keys()).index(use_case_default),
        key="bess_use_case",
    )
    st.markdown(f"<div class='mode-badge'>{DESCRIPCIONES_USE_CASE[use_case]}</div>",
                unsafe_allow_html=True)

    df_metricas = calcular_metricas_bess(df)
    if df_metricas.empty:
        st.warning("No hay suficientes datos para calcular métricas BESS.")
        return

    pesos_default = PESOS_BESS[use_case].copy()
    pesos = pesos_default.copy()
    with st.expander("⚙️ Ajustar pesos del scoring (opcional)"):
        st.caption("Los pesos se normalizarán automáticamente.")
        col_p = st.columns(min(3, len(pesos_default)))
        for i, (metrica, valor) in enumerate(pesos_default.items()):
            with col_p[i % len(col_p)]:
                pesos[metrica] = st.slider(
                    metrica.replace('_', ' ').title(),
                    min_value=0.0, max_value=1.0, value=valor, step=0.05,
                    key=f"peso_{use_case}_{metrica}",
                )
        suma = sum(pesos.values())
        if abs(suma - 1.0) > 0.05:
            st.caption(f"⚠️ Pesos suman {suma:.2f}, se normalizará a 1.0")
            pesos = {k: v/suma for k, v in pesos.items()} if suma > 0 else pesos_default

    df_score = calcular_score_bess(df_metricas, use_case, pesos=pesos)

    st.markdown("### 🏆 Top 3 nodos recomendados")
    top3 = df_score.head(3)
    cols_t = st.columns(3)
    medals = ['🥇', '🥈', '🥉']
    sym = label_moneda_short()
    for i, (idx, row) in enumerate(top3.iterrows()):
        with cols_t[i]:
            st.metric(
                f"{medals[i]} {row['nodo']}",
                f"{row['score']:.1f}",
                help=(f"PML promedio: {sym}{row['pml_promedio']:.2f} | "
                      f"Volatilidad: {sym}{row['volatilidad']:.2f} | "
                      f"Spread P95-P5: {sym}{row['spread_p95_p5']:.2f}")
            )

    fig_rank = grafica_bess_ranking(df_score, use_case, top_n=min(15, len(df_score)))
    if fig_rank: st.plotly_chart(fig_rank, use_container_width=True)

    st.markdown("### 📋 Tabla completa con métricas")
    st.dataframe(
        df_score, use_container_width=True, hide_index=True,
        column_config={
            "nodo":           st.column_config.TextColumn("Clave", width="small"),
            "score":          st.column_config.NumberColumn("Score BESS", format="%.1f"),
            "pml_promedio":   st.column_config.NumberColumn("PML promedio", format=fmt_moneda()),
            "volatilidad":    st.column_config.NumberColumn("Volatilidad", format=fmt_moneda()),
            "spread_p95_p5":  st.column_config.NumberColumn("Spread P95-P5", format=fmt_moneda()),
            "spread_avg_diario": st.column_config.NumberColumn("Spread día prom", format=fmt_moneda()),
            "spread_dia":     st.column_config.NumberColumn("Spread día/noche", format=fmt_moneda()),
            "cambios_bruscos": st.column_config.NumberColumn("Cambios bruscos", format="%d"),
            "horas_pico":     st.column_config.NumberColumn("Horas pico", format="%d"),
            "pct_horas_neg":  st.column_config.NumberColumn("% horas neg", format="%.1f%%"),
        },
    )


# ═══════════════════════════════════════════════════════════════════════
# RENDER DASHBOARD COMPLETO (sin Excel - solo análisis)
# ═══════════════════════════════════════════════════════════════════════
def render_dashboard(acumulado, catalogo, matches_df=None):
    st.divider()
    st.markdown("## 📊 Dashboard analítico")
    st.caption("Pasa el mouse sobre los gráficos para ver valores específicos.")

    df = acumulado_a_dataframe(acumulado, catalogo)
    if df.empty:
        st.warning("No hay datos numéricos válidos.")
        return

    col1, col2, col3, col4 = st.columns(4)
    pml_global = df["pml"].mean()
    pml_max = df["pml"].max()
    pct_neg = (df["pml"] < 0).sum() / len(df) * 100
    nodo_top = df.groupby("nodo")["pml"].mean().idxmax()
    sym = label_moneda_short()
    col1.metric("PML promedio global", f"{sym}{pml_global:.2f}")
    col2.metric("PML máximo", f"{sym}{pml_max:.2f}")
    col3.metric("% horas negativas", f"{pct_neg:.1f}%")
    col4.metric("Nodo top", nodo_top)

    # MAPA
    if matches_df is not None and not matches_df.empty:
        st.markdown("### 🗺️ Mapa interactivo de nodos")
        n_mapeados = matches_df['lat'].notna().sum()
        n_consultados = len(matches_df)
        if n_mapeados == 0:
            st.warning(f"❌ Ninguno de los {n_consultados} nodos pudo ser mapeado en OSM.")
        else:
            cm1, cm2, cm3 = st.columns(3)
            cm1.metric("Mapeados", f"{n_mapeados}/{n_consultados}")
            cc = matches_df['calidad'].value_counts()
            n_exc = cc.get('🥇 Excelente', 0)
            n_bue = cc.get('🥈 Bueno', 0)
            n_ace = cc.get('🥉 Aceptable', 0)
            cm2.metric("🥇 Excelente / 🥈 Bueno", f"{n_exc} / {n_bue}")
            cm3.metric("🥉 Aceptable", n_ace)

            fig_mapa = grafica_mapa(matches_df, df, color_by='pml')
            if fig_mapa:
                st.plotly_chart(fig_mapa, use_container_width=True,
                                config={'scrollZoom': True, 'displayModeBar': True})

            # Leyenda CCR
            render_leyenda_ccr(matches_df)

            render_panel_ccr(matches_df, df)

            sin_match = matches_df[matches_df['lat'].isna()]
            if not sin_match.empty:
                with st.expander(f"❌ {len(sin_match)} nodos sin match"):
                    st.dataframe(
                        sin_match[['clave', 'nombre_cenace', 'ccr', 'estado', 'razon']],
                        use_container_width=True, hide_index=True,
                    )

    # TABLA RESUMEN
    st.markdown("### 📋 Tabla resumen por nodo")
    summary = calcular_resumen(df)
    if not summary.empty:
        st.dataframe(
            summary, use_container_width=True, hide_index=True,
            column_config={
                "nodo":        st.column_config.TextColumn("Clave", width="small"),
                "nombre":      st.column_config.TextColumn("Nombre", width="medium"),
                "ccr":         st.column_config.TextColumn("CCR", width="small"),
                "registros":   st.column_config.NumberColumn("# Reg", format="%d"),
                "promedio":    st.column_config.NumberColumn("Promedio", format=fmt_moneda()),
                "mediana":     st.column_config.NumberColumn("Mediana", format=fmt_moneda()),
                "maximo":      st.column_config.NumberColumn("Máximo", format=fmt_moneda()),
                "minimo":      st.column_config.NumberColumn("Mínimo", format=fmt_moneda()),
                "std":         st.column_config.NumberColumn("Volatilidad", format=fmt_moneda()),
                "p95":         st.column_config.NumberColumn("P95", format=fmt_moneda()),
                "p5":          st.column_config.NumberColumn("P5", format=fmt_moneda()),
                "% horas neg": st.column_config.NumberColumn("% Neg", format="%.1f%%"),
            },
        )

    st.markdown("### 📈 Evolución temporal del PML")
    fig_lin = grafica_lineas_tiempo(df, max_nodos=min(15, len(acumulado)))
    if fig_lin: st.plotly_chart(fig_lin, use_container_width=True)

    st.markdown("### 🔥 Heatmap horario × mensual")
    nodos_disp = sorted(df["nodo"].unique())
    nodo_h = st.selectbox("Selecciona nodo:", nodos_disp, index=0, key="sel_heatmap")
    fig_heat = grafica_heatmap_horario(df, nodo_h)
    if fig_heat: st.plotly_chart(fig_heat, use_container_width=True)

    # ─── PML PROMEDIO MULTI-AÑO ───
    años_total = df['fecha_dt'].dt.year.nunique() if 'fecha_dt' in df.columns else 1
    st.markdown("### 📊 PML promedio mensual por año")
    if años_total < 1:
        st.caption("Sin datos suficientes para gráfica multi-año.")
    else:
        nodo_multi = st.selectbox(
            "Selecciona nodo para ver evolución mensual por año:",
            nodos_disp,
            index=nodos_disp.index(nodo_h) if nodo_h in nodos_disp else 0,
            key="sel_multiano",
            help="Muestra promedio mensual del PML del nodo, comparando años disponibles."
        )
        fig_multi, años_disp = grafica_pml_multiano(df, nodo_multi)
        if fig_multi:
            st.plotly_chart(fig_multi, use_container_width=True)
            if len(años_disp) == 1:
                st.caption(f"ℹ️ Solo hay datos del año {años_disp[0]}. "
                           f"Para comparar años, consulta un período que cubra más de un año natural.")
        else:
            st.warning("No se pudo generar la gráfica multi-año.")

    # ─── SPREAD POR CCR (solo si ≥2 CCRs distintos) ───
    ccrs_unicos_dash = sorted([c for c in df['ccr'].unique() if c and c != "?"])
    if len(ccrs_unicos_dash) >= 2:
        st.markdown("### 📊 Spread mensual por CCR (zonas)")
        st.caption(
            f"Comparativa continua del PML promedio mensual entre **{len(ccrs_unicos_dash)} CCRs** "
            f"presentes en tu consulta. Útil para identificar diferenciales entre zonas."
        )
        fig_spread, _ = grafica_spread_ccr(df)
        if fig_spread:
            st.plotly_chart(fig_spread, use_container_width=True)

    st.markdown("### 🏆 Rankings comparativos")
    tab1, tab2, tab3 = st.tabs(["Mayor PML", "Mayor volatilidad", "Más horas negativas"])
    with tab1:
        fig = grafica_barras_top(df, "promedio", top_n=min(10, len(acumulado)))
        if fig: st.plotly_chart(fig, use_container_width=True)
    with tab2:
        fig = grafica_barras_top(df, "volatilidad", top_n=min(10, len(acumulado)))
        if fig: st.plotly_chart(fig, use_container_width=True)
    with tab3:
        fig = grafica_barras_top(df, "negativos", top_n=min(10, len(acumulado)))
        if fig: st.plotly_chart(fig, use_container_width=True)

    # BESS scoring
    render_bess_scoring(df)


# ═══════════════════════════════════════════════════════════════════════
# RENDER CENTRO DE DESCARGAS (Modo Solo Datos)
# ═══════════════════════════════════════════════════════════════════════
def render_centro_descargas(acumulado, catalogo, sistema, proceso, fecha_ini, fecha_fin,
                              matches_df=None, moneda="MXN", tc_info=None):
    st.divider()
    st.markdown("## 📥 Centro de descargas")
    moneda_label = f"💱 Análisis en **{moneda}**"
    if moneda == "USD" and tc_info and tc_info.get("tc_promedio"):
        moneda_label += f" · TC FIX promedio: **{tc_info['tc_promedio']:.4f} MXN/USD**"
    st.caption(f"Selecciona los archivos que necesites. {moneda_label}")
    st.caption("⚡ Los archivos solo se generan cuando das click en cada botón.")

    df = acumulado_a_dataframe(acumulado, catalogo)
    df_resumen = calcular_resumen(df) if not df.empty else pd.DataFrame()
    df_metricas = calcular_metricas_bess(df) if not df.empty else pd.DataFrame()
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    sufijo = f"_{moneda}"
    n_nodos = len(acumulado)

    # ─── Multi-año (todos los nodos si ≤20) ───
    df_multianos_dict = None  # dict {nodo: df_pivot}
    incluir_multianos = False

    if n_nodos > 0 and n_nodos <= 20:
        st.markdown("##### 📊 Gráficas multi-año en Excel de Análisis")
        st.caption(
            f"Como tu consulta tiene **{n_nodos} nodos** (≤20), el Excel de Análisis "
            f"incluirá una hoja con gráfica nativa para **CADA nodo** mostrando el PML "
            f"promedio mensual por año. Total: **{n_nodos} hojas con gráficas**."
        )
        incluir_multianos = st.toggle(
            f"Incluir gráficas multi-año de los {n_nodos} nodos",
            value=True,
            key="tog_incluir_multianos",
            help="Una hoja por nodo con tabla + gráfica de líneas nativa de Excel"
        )

        if incluir_multianos and not df.empty:
            df_multianos_dict = {}
            for nodo_n in sorted(df["nodo"].unique()):
                sub = df[df["nodo"] == nodo_n].copy()
                if sub.empty:
                    continue
                sub["año"] = sub["fecha_dt"].dt.year
                sub["mes_num"] = sub["fecha_dt"].dt.month
                pivot = sub.pivot_table(values="pml", index="mes_num",
                                          columns="año", aggfunc="mean").round(2)
                df_multianos_dict[str(nodo_n)] = pivot.reset_index()

            # Mostrar resumen de años disponibles
            sample_keys = list(df_multianos_dict.keys())[:1]
            if sample_keys:
                años_count = len([c for c in df_multianos_dict[sample_keys[0]].columns
                                   if c != 'mes_num'])
                if años_count >= 2:
                    st.success(
                        f"✅ {n_nodos} gráficas con **{años_count} años cada una** listas "
                        f"para incluirse en el Excel de Análisis."
                    )
                else:
                    st.info(
                        f"ℹ️ Solo hay datos de **1 año** en la consulta. La gráfica "
                        f"multi-año se incluirá pero tendrá una sola línea. Para "
                        f"comparar años, consulta un período más largo."
                    )

    elif n_nodos > 20:
        st.info(
            f"⚠️ Tu consulta tiene **{n_nodos} nodos**. Las gráficas multi-año "
            f"solo se incluyen automáticamente si hay ≤20 nodos (para mantener el "
            f"Excel manejable). Reduce tu consulta si quieres incluirlas."
        )

    st.markdown("##### 📦 Tipos de archivos disponibles")

    # ─── Descripciones contextuales según tamaño ───
    if n_nodos < 20:
        desc_datos = f"📊 Excel **ligero** (~{n_nodos*150} KB · ~5s) — portada + resumen + 1 hoja por nodo"
        desc_anal  = f"📈 Excel **compacto** (~80 KB · ~3s) — BESS Scoring 3 use cases + métricas"
        if incluir_multianos:
            desc_anal += " + gráfica multi-año"
    elif n_nodos <= 50:
        desc_datos = f"📊 Excel **medio** (~{n_nodos*200} KB · ~30s) — 1 hoja por nodo"
        desc_anal  = f"📈 Excel **compacto** (~100 KB · ~5s) — BESS Scoring 3 use cases + métricas"
    elif n_nodos <= 100:
        desc_datos = f"📊 Excel **pesado** (~{n_nodos*250} KB · 1-2 min) — 1 hoja por nodo · puede tardar al abrir"
        desc_anal  = f"📈 Excel **compacto** (~120 KB · ~10s) — BESS Scoring 3 use cases + métricas"
    else:
        desc_datos = f"📊 Excel **muy pesado** (~{n_nodos*300} KB · 2-5 min) — ⚠️ usa modo Solo Datos con cuidado"
        desc_anal  = f"📈 Excel **compacto** (~150 KB · ~20s) — BESS Scoring 3 use cases + métricas"

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("##### 📊 Excel de Datos")
        st.caption(desc_datos)
        if "excel_datos_bytes" not in st.session_state:
            st.session_state["excel_datos_bytes"] = None
        if st.button("Generar y descargar", key="btn_gen_datos", type="primary",
                      use_container_width=True):
            with st.spinner(f"Generando Excel de datos ({n_nodos} nodos)..."):
                st.session_state["excel_datos_bytes"] = generar_excel_datos(
                    acumulado, sistema, proceso, fecha_ini, fecha_fin,
                    moneda=moneda, tc_info=tc_info)
        if st.session_state["excel_datos_bytes"]:
            st.download_button(
                label="📥 Descargar archivo",
                data=st.session_state["excel_datos_bytes"],
                file_name=f"PML_CENACE_Datos{sufijo}_{sistema}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_datos",
                use_container_width=True,
            )

    with col2:
        st.markdown("##### 📈 Excel de Análisis")
        st.caption(desc_anal)
        if "excel_analisis_bytes" not in st.session_state:
            st.session_state["excel_analisis_bytes"] = None
        if st.button("Generar y descargar", key="btn_gen_anal", type="primary",
                      use_container_width=True):
            if df_metricas.empty:
                st.error("No hay suficientes datos para generar análisis.")
            else:
                with st.spinner("Generando Excel de análisis..."):
                    st.session_state["excel_analisis_bytes"] = generar_excel_analisis(
                        df_metricas, df_resumen, sistema, proceso, fecha_ini, fecha_fin,
                        moneda=moneda, tc_info=tc_info,
                        df_multianos=df_multianos_dict if incluir_multianos else None)
        if st.session_state["excel_analisis_bytes"]:
            st.download_button(
                label="📥 Descargar archivo",
                data=st.session_state["excel_analisis_bytes"],
                file_name=f"PML_CENACE_Analisis_BESS{sufijo}_{sistema}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_anal",
                use_container_width=True,
            )

    with col3:
        st.markdown("##### 🌍 KMZ Geográfico")
        st.caption("🌍 Archivo Google Earth (~10-200 KB) — nodos mapeados con calidad del match")
        if matches_df is None or matches_df.empty:
            st.info("⚠️ La geocodificación no se hizo. Activa el toggle '🗺️ Incluir geocodificación' en el sidebar y vuelve a ejecutar.")
        else:
            n_mapeados = matches_df['lat'].notna().sum()
            if n_mapeados == 0:
                st.warning("Ningún nodo pudo ser mapeado.")
            else:
                if "kmz_bytes" not in st.session_state:
                    st.session_state["kmz_bytes"] = None
                if st.button(f"Generar y descargar ({n_mapeados} nodos)",
                              key="btn_gen_kmz", type="primary",
                              use_container_width=True):
                    with st.spinner("Generando KMZ..."):
                        st.session_state["kmz_bytes"] = generar_kmz(matches_df)
                if st.session_state["kmz_bytes"]:
                    st.download_button(
                        label="📥 Descargar archivo",
                        data=st.session_state["kmz_bytes"],
                        file_name=f"PML_CENACE_Geo_{ts}.kmz",
                        mime="application/vnd.google-earth.kmz",
                        key="dl_kmz",
                        use_container_width=True,
                    )

    # ─── EXCEL CUSTOM ───
    st.divider()
    st.markdown("##### 🎨 Excel Custom — descarga personalizada")
    st.caption(
        "Selecciona qué secciones incluir en un Excel a la medida. "
        "Cada toggle agrega o quita una hoja del archivo final."
    )

    # Detectar disponibilidad de cada opción
    n_nodos_actual = df["nodo"].nunique() if not df.empty else 0
    ccrs_unicos_cust = sorted([c for c in df['ccr'].unique() if c and c != "?"]) if not df.empty else []
    multiano_disponible = (n_nodos_actual > 0 and n_nodos_actual <= 20)

    cc1, cc2, cc3 = st.columns(3)
    with cc1:
        opt_resumen = st.toggle("📋 Resumen estadístico", value=True, key="cust_resumen")
        opt_top_pml = st.toggle("🏆 Top PML promedio", value=False, key="cust_top_pml")
        opt_top_vol = st.toggle("📊 Top volatilidad", value=False, key="cust_top_vol")
    with cc2:
        opt_arb = st.toggle("💰 BESS Arbitraje", value=False, key="cust_arb")
        opt_ssaa = st.toggle("⚙️ BESS SSAA", value=False, key="cust_ssaa")
        opt_firm = st.toggle("🌅 BESS Firming", value=False, key="cust_firm")
    with cc3:
        opt_multi = st.toggle(
            f"📈 Multi-año ({n_nodos_actual} hojas)" if multiano_disponible else "📈 Multi-año (>20 nodos no)",
            value=False, disabled=not multiano_disponible, key="cust_multi"
        )
        spread_label = (f"📊 Spread CCR ({len(ccrs_unicos_cust)} zonas)"
                         if len(ccrs_unicos_cust) >= 2 else "📊 Spread CCR (necesita ≥2 CCRs)")
        opt_spread = st.toggle(spread_label, value=False,
                                disabled=len(ccrs_unicos_cust) < 2, key="cust_spread")

    opciones_custom = {
        "resumen": opt_resumen,
        "multiano": opt_multi,
        "spread_ccr": opt_spread,
        "bess_arbitraje": opt_arb,
        "bess_ssaa": opt_ssaa,
        "bess_firming": opt_firm,
        "top_pml": opt_top_pml,
        "top_volatilidad": opt_top_vol,
    }
    n_opciones = sum(1 for v in opciones_custom.values() if v)

    if n_opciones == 0:
        st.info("ℹ️ Selecciona al menos una sección arriba para generar el Excel Custom.")
    else:
        st.caption(f"✅ **{n_opciones} secciones** seleccionadas para el Excel Custom.")

        # Construir df_multianos_dict si está activo
        df_multi_for_custom = None
        if opt_multi and multiano_disponible:
            df_multi_for_custom = {}
            for nodo_n in sorted(df["nodo"].unique()):
                sub = df[df["nodo"] == nodo_n].copy()
                if sub.empty:
                    continue
                sub["año"] = sub["fecha_dt"].dt.year
                sub["mes_num"] = sub["fecha_dt"].dt.month
                pivot = sub.pivot_table(values="pml", index="mes_num",
                                          columns="año", aggfunc="mean").round(2)
                df_multi_for_custom[str(nodo_n)] = pivot.reset_index()

        if "excel_custom_bytes" not in st.session_state:
            st.session_state["excel_custom_bytes"] = None

        if st.button("🎨 Generar Excel Custom", key="btn_gen_custom",
                      type="primary", use_container_width=True):
            with st.spinner(f"Generando Excel Custom con {n_opciones} secciones..."):
                st.session_state["excel_custom_bytes"] = generar_excel_custom(
                    df, df_resumen, df_metricas, opciones_custom,
                    sistema, proceso, fecha_ini, fecha_fin,
                    moneda=moneda, tc_info=tc_info,
                    df_multianos_dict=df_multi_for_custom,
                )

        if st.session_state["excel_custom_bytes"]:
            st.download_button(
                label="📥 Descargar Excel Custom",
                data=st.session_state["excel_custom_bytes"],
                file_name=f"PML_CENACE_Custom{sufijo}_{sistema}_{ts}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_custom",
                use_container_width=True,
            )


# ═══════════════════════════════════════════════════════════════════════
# UI PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="main-header">
    <h1>⚡ CENACE PML Analyzer 
        <span class="srf-badge">SRF</span>
    </h1>
    <p>Descarga, análisis, mapeo y BESS scoring · MXN o USD (Banxico FIX) — Recurrent Energy / Canadian Solar</p>
</div>
""", unsafe_allow_html=True)

catalogo, catalogo_fuente = get_catalogo_activo()
if catalogo:
    badge_fuente = "📦 default" if catalogo_fuente == "repositorio (data/catalogo_nodos.xlsx)" else "📤 actualizado"
    st.info(f"📚 Catálogo CENACE cargado: **{len(catalogo):,} nodos** disponibles · "
            f"{badge_fuente}")
else:
    st.warning("⚠️ Catálogo no cargado.")

# Inicializar session_state
if "consulta_ejecutada" not in st.session_state:
    st.session_state["consulta_ejecutada"] = False
if "acumulado" not in st.session_state:
    st.session_state["acumulado"] = {}
if "matches_df" not in st.session_state:
    st.session_state["matches_df"] = None
if "consulta_params" not in st.session_state:
    st.session_state["consulta_params"] = {}

with st.sidebar:
    st.markdown(
        f"<h2 style='color:{RE_BLUE};margin-bottom:0;font-weight:700;'>"
        f"⚡🔋 Node Settings</h2>",
        unsafe_allow_html=True,
    )
    st.caption("Configuración")

    st.markdown("**🚀 Modo de uso**")
    modo = st.radio(
        "Selecciona qué hacer:",
        options=["🔬 Análisis", "📊 Datos", "🗺️ Mapa"],
        index=0,
        help=(
            "**Análisis**: descarga + dashboard + mapa + BESS scoring (sin Excel)\n\n"
            "**Datos**: descarga + Centro de descargas (Excel/KMZ)\n\n"
            "**Mapa**: solo geocodifica (~5 segundos)"
        ),
        label_visibility="collapsed",
    )

    es_solo_mapa = (modo == "🗺️ Mapa")
    es_solo_datos = (modo == "📊 Datos")
    es_completo = (modo == "🔬 Análisis")
    necesita_datos = es_solo_datos or es_completo
    necesita_mapa = es_solo_mapa or es_completo

    # FX Selector — solo si se usan datos
    if necesita_datos:
        st.markdown("---")
        st.markdown("**💱 FX Selector**")
        moneda_sel = st.radio(
            "Moneda:",
            options=["🇲🇽 MXN (pesos)", "🇺🇸 USD (dólares)"],
            index=0,
            label_visibility="collapsed",
            help=(
                "**MXN**: PML directo del CENACE en pesos.\n\n"
                "**USD**: convertido día por día con el tipo de cambio "
                "FIX de Banxico (serie SF43718)."
            ),
        )
        st.session_state["moneda"] = "USD" if "USD" in moneda_sel else "MXN"
    else:
        st.session_state["moneda"] = "MXN"

    st.markdown("---")
    st.markdown("**📍 Nodos CENACE**")

    # Procesar pending del botón "Agregar al textarea" (debe ir ANTES del text_area)
    if "nodos_pending" in st.session_state and st.session_state["nodos_pending"]:
        # Mover de pending al input
        st.session_state["nodos_input"] = st.session_state["nodos_pending"]
        st.session_state["nodos_pending"] = ""

    nodos_input = st.text_area(
        "Lista de claves",
        height=120,
        placeholder="01VAJ-230\n01XAL-230\n06FUN-115",
        key="nodos_input",
    )

    # Badge si los nodos vinieron del filtro
    if st.session_state.get("nodos_from_filter"):
        st.caption(f"📍 **{st.session_state['nodos_from_filter']} nodos del filtro**")

    # ─── SELECTOR JERÁRQUICO ───
    with st.expander("🔍 Buscar nodos por filtros (Estado / Zona / CCR / kV)"):
        if catalogo:
            cat_df = pd.DataFrame.from_dict(catalogo, orient='index').reset_index()
            cat_df.rename(columns={'index': 'clave'}, inplace=True)

            sistemas_disp = sorted(cat_df['sistema'].dropna().unique())
            sistemas_disp = [s for s in sistemas_disp if s]
            sis_sel = st.selectbox("🌎 Sistema", ["Todos"] + sistemas_disp, key="filt_sis")
            df_f = cat_df.copy() if sis_sel == "Todos" else cat_df[cat_df['sistema'] == sis_sel]

            estados_disp = sorted([e for e in df_f['estado'].dropna().unique() if e])
            est_sel = st.multiselect("🏛️ Estado(s)", estados_disp, key="filt_est")
            if est_sel:
                df_f = df_f[df_f['estado'].isin(est_sel)]

            municipios_disp = sorted([m for m in df_f['municipio'].dropna().unique() if m])
            mun_sel = st.multiselect("🏘️ Municipio(s)", municipios_disp, key="filt_mun")
            if mun_sel:
                df_f = df_f[df_f['municipio'].isin(mun_sel)]

            zonas_disp = sorted([z for z in df_f['zona'].dropna().unique() if z])
            zona_sel = st.multiselect("🗺️ Zona CENACE", zonas_disp, key="filt_zona")
            if zona_sel:
                df_f = df_f[df_f['zona'].isin(zona_sel)]

            ccr_disp = sorted([c for c in df_f['ccr'].dropna().unique() if c])
            ccr_sel = st.multiselect("📍 CCR (Centro Control Regional)", ccr_disp, key="filt_ccr")
            if ccr_sel:
                df_f = df_f[df_f['ccr'].isin(ccr_sel)]

            kvs_disp = sorted([int(k) for k in df_f['kv'].dropna().unique() if k > 0])
            kv_sel = st.multiselect("⚡ Voltaje (kV)", kvs_disp, key="filt_kv")
            if kv_sel:
                df_f = df_f[df_f['kv'].isin(kv_sel)]

            st.markdown(f"**📋 {len(df_f):,} nodos coinciden**")

            if len(df_f) > 0:
                # Buscador opcional
                buscar = st.text_input("🔎 Buscar por nombre/clave (opcional)", key="filt_buscar")
                if buscar:
                    mask = (df_f['clave'].str.contains(buscar, case=False, na=False) |
                            df_f['nombre'].str.contains(buscar, case=False, na=False))
                    df_f = df_f[mask]
                    st.caption(f"Tras filtro: **{len(df_f):,} nodos**")

                # Mostrar tabla compacta
                if len(df_f) > 0:
                    st.dataframe(
                        df_f[['clave', 'nombre', 'ccr', 'estado', 'kv']].head(100),
                        use_container_width=True, hide_index=True,
                        column_config={
                            "clave":  st.column_config.TextColumn("Clave", width="small"),
                            "nombre": st.column_config.TextColumn("Nombre", width="medium"),
                            "ccr":    st.column_config.TextColumn("CCR", width="small"),
                            "estado": st.column_config.TextColumn("Estado", width="small"),
                            "kv":     st.column_config.NumberColumn("kV", format="%d"),
                        },
                        height=200,
                    )
                    if len(df_f) > 100:
                        st.caption(f"⚠️ Mostrando solo los primeros 100 (refina filtros para ver todos)")

                    # Botón para agregar al textarea — limpia consulta previa
                    if st.button(f"✅ Agregar {len(df_f)} nodos (reemplaza textarea)",
                                 key="btn_add_nodos"):
                        nuevas_claves = "\n".join(df_f['clave'].tolist())
                        # Reemplazar (no concatenar) para que sea predecible
                        st.session_state["nodos_pending"] = nuevas_claves
                        st.session_state["nodos_from_filter"] = len(df_f)
                        # Limpiar resultado de consulta anterior (forzar re-ejecutar)
                        for k in ("consulta_ejecutada", "acumulado", "matches_df",
                                  "fx_info", "consulta_params",
                                  "excel_datos_bytes", "excel_analisis_bytes",
                                  "excel_custom_bytes", "kmz_bytes"):
                            if k in st.session_state:
                                if k == "consulta_ejecutada":
                                    st.session_state[k] = False
                                else:
                                    st.session_state[k] = None
                        st.rerun()
        else:
            st.warning("No hay catálogo cargado para filtrar.")

    if necesita_datos:
        # Solo proceso MTR/MDA — Sistema se auto-detecta del catálogo
        proceso = st.selectbox("⚙️ Proceso", ["MTR", "MDA"], index=0, key="proceso_sel",
                                 help="**MTR** = Mercado en Tiempo Real (precios reales). "
                                      "**MDA** = Mercado del Día en Adelante (pronosticados).")
        st.caption("ℹ️ Sistema (SIN/BCA/BCS) se auto-detecta de cada nodo desde el catálogo.")

        st.markdown("**📅 Período**")
        col_f1, col_f2 = st.columns(2)
        today = date.today()
        # Default: hoy −2 semanas (delay típico CENACE) hasta hoy −1 día
        with col_f1:
            f_ini = st.date_input("Desde", value=today - timedelta(days=90),
                                    max_value=today, key="f_ini")
        with col_f2:
            f_fin = st.date_input("Hasta", value=today - timedelta(days=14),
                                    max_value=today, key="f_fin",
                                    help="CENACE típicamente tiene delay de 1-2 semanas.")

        # Warning si fin > hoy −7 días
        if f_fin > today - timedelta(days=7):
            st.caption(
                "⚠️ La fecha fin es muy reciente. CENACE suele publicar con delay de 1-2 semanas. "
                "Podrías recibir datos parciales."
            )
        sistema = None  # auto-detect
    else:
        proceso = "MTR"
        sistema = None
        today = date.today()
        f_ini = today - timedelta(days=90)
        f_fin = today - timedelta(days=14)

    # Toggle de geocodificación en modo solo datos
    if es_solo_datos:
        st.markdown("**🗺️ Geocodificación**")
        incluir_geo_solo_datos = st.toggle("Incluir geocodificación",
                                            value=False,
                                            key="toggle_geo_datos",
                                            help="Activar si vas a descargar el KMZ. Tarda ~5s extra.")
    else:
        incluir_geo_solo_datos = False

    st.divider()

    # Botón limpiar todo
    if st.button("🧹 Limpiar todo", key="btn_clear_all", use_container_width=True,
                  help="Reset completo: borra textarea, filtros, datos descargados y descargas pendientes."):
        # Limpiar todo el session_state
        keys_to_clear = [
            "consulta_ejecutada", "acumulado", "matches_df", "fx_info",
            "consulta_params", "nodos_input", "nodos_pending", "nodos_from_filter",
            "excel_datos_bytes", "excel_analisis_bytes", "excel_custom_bytes",
            "kmz_bytes", "moneda",
            # Filtros
            "filt_sis", "filt_est", "filt_mun", "filt_zona", "filt_ccr",
            "filt_kv", "filt_buscar",
            # Selectores
            "sel_heatmap", "sel_multiano", "sel_multianos_excel",
            "tog_incluir_multianos", "conf_grande",
        ]
        for k in keys_to_clear:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()

    with st.expander("📤 Actualizar catálogo de nodos"):
        st.caption(
            "El catálogo se actualiza ~1 vez al mes. Si tienes una versión más nueva, "
            "súbelo aquí (.xlsx) y reemplazará el catálogo durante esta sesión."
        )
        uploaded_cat = st.file_uploader(
            "Catálogo CENACE (.xlsx)",
            type=["xlsx"],
            key="cat_upload",
            help="Archivo `Catálogo_NodosP_Sistema_Eléctrico_Nacional` del portal CENACE",
        )

        # Flag para evitar reprocesar el mismo archivo en cada rerun
        if uploaded_cat is not None:
            file_id = f"{uploaded_cat.name}_{uploaded_cat.size}"
            last_processed = st.session_state.get("cat_last_processed", "")
            if file_id != last_processed:
                # Solo procesar si es un archivo distinto al ya procesado
                file_bytes = uploaded_cat.read()
                new_cat = cargar_catalogo_uploaded(file_bytes)
                if new_cat:
                    st.session_state["catalogo_uploaded"] = new_cat
                    st.session_state["cat_last_processed"] = file_id
                    st.success(
                        f"✅ Catálogo actualizado: **{len(new_cat):,} nodos**. "
                        f"Aplica desde la próxima interacción."
                    )
                else:
                    st.error("❌ No se pudo leer el archivo. Verifica que sea el catálogo CENACE oficial.")
                    st.session_state["cat_last_processed"] = file_id  # marcar como procesado para no reintentar
            else:
                # Ya procesado: solo mostrar el estado actual
                if st.session_state.get("catalogo_uploaded"):
                    st.success(
                        f"✅ Catálogo actualizado en esta sesión: "
                        f"**{len(st.session_state['catalogo_uploaded']):,} nodos**"
                    )

        if "catalogo_uploaded" in st.session_state and st.session_state["catalogo_uploaded"]:
            if st.button("🔄 Volver al catálogo default", key="btn_reset_cat"):
                st.session_state["catalogo_uploaded"] = None
                st.session_state["cat_last_processed"] = ""
                st.rerun()

    st.caption("Sebastian Roldan (SRF)\nRecurrent Energy · Canadian Solar")


col_left, col_right = st.columns([2, 1])
with col_left:
    st.markdown("### 🚀 Ejecutar consulta")

    modo_descripcion = {
        "🔬 Análisis": "Descargará datos PML + mapeará nodos + dashboard + BESS scoring (sin Excel).",
        "📊 Datos": "Descargará datos PML y mostrará un Centro de Descargas con Excel/KMZ.",
        "🗺️ Mapa": ("Solo localizará los nodos en OpenStreetMap. <b>Sin descargar PML.</b>"),
    }
    st.markdown(
        f"<div class='mode-badge'><b>Modo: {modo}</b><br>{modo_descripcion[modo]}</div>",
        unsafe_allow_html=True,
    )

    nodos = []
    if nodos_input.strip():
        nodos_raw = re.split(r"[,\n]+", nodos_input.strip())
        nodos = [n.strip().upper() for n in nodos_raw if n.strip()]
        seen = set()
        nodos = [n for n in nodos if not (n in seen or seen.add(n))]

    if nodos:
        cv = [n for n in nodos if n in catalogo] if catalogo else nodos
        ci = [n for n in nodos if catalogo and n not in catalogo]
        c_a, c_b, c_c = st.columns(3)
        c_a.metric("Solicitados", len(nodos))
        c_b.metric("✅ En catálogo", len(cv))
        c_c.metric("⚠️ Desconocidos", len(ci))
        if ci:
            with st.expander(f"⚠️ {len(ci)} nodos no encontrados"):
                st.code("\n".join(ci))

    if nodos and necesita_datos and f_ini < f_fin:
        n_dias = (f_fin - f_ini).days + 1
        n_bloques = (n_dias + BLOQUE_MAX - 1) // BLOQUE_MAX
        n_lotes = (len(nodos) + 9) // 10
        n_consultas = n_bloques * n_lotes
        tiempo = n_consultas * 1.5 / MAX_WORKERS
        st.caption(f"📊 **{n_dias} días · {n_consultas} consultas · ~{tiempo/60:.1f} min**")
    elif nodos and es_solo_mapa:
        st.caption(f"🗺️ Solo geocodificación de **{len(nodos)} nodos** (~5 segundos)")

    # ─── WARNINGS PROGRESIVOS POR # NODOS ───
    n_nodos = len(nodos)
    confirmacion_grande = True
    bloqueo_duro = False

    if n_nodos > 0:
        if es_solo_mapa:
            # Modo Mapa permite hasta 2500 (es ligero)
            if n_nodos > 2500:
                st.error(
                    f"🔴 **{n_nodos} nodos** es demasiado para procesar incluso en modo Mapa. "
                    f"Por estabilidad, sugiero dividir en consultas menores a 2500 nodos."
                )
                bloqueo_duro = True
            elif n_nodos > 500:
                st.warning(
                    f"⚠️ **{n_nodos} nodos** es una consulta grande. "
                    f"Modo Mapa puede tomar 30-60s. Si quieres datos PML, divide en partes."
                )
        else:
            # Modo Análisis o Datos: límites más estrictos
            if n_nodos > 300:
                st.error(
                    f"🔴 **{n_nodos} nodos** excede el límite recomendado para Análisis/Datos. "
                    f"La app web puede colgarse por OOM (limite 1GB RAM en Streamlit Cloud Free). "
                    f"Sugiero dividir en consultas de 100-200 nodos. "
                    f"💡 Si solo quieres ver dónde están en el mapa, usa modo **🗺️ Mapa** "
                    f"(soporta hasta 2500 nodos)."
                )
                bloqueo_duro = True
            elif n_nodos >= 100:
                st.warning(
                    f"🟠 **{n_nodos} nodos** es una consulta grande. "
                    f"Puede tomar 5-10 minutos y consumir mucha memoria."
                )
                confirmacion_grande = st.checkbox(
                    f"He revisado y quiero proceder con {n_nodos} nodos",
                    key="conf_grande"
                )
            elif n_nodos >= 30:
                st.info(
                    f"⚠️ **{n_nodos} nodos** — esto puede tomar 2-5 minutos."
                )

    boton = st.button(
        "⚡ Ejecutar",
        type="primary",
        disabled=(len(nodos) == 0 or bloqueo_duro or not confirmacion_grande)
    )

with col_right:
    st.markdown("### 📋 Instrucciones")
    if es_solo_mapa:
        st.markdown("""
        1. Pega claves de nodos
        2. Click en **Ejecutar**
        3. Ve la ubicación geográfica en el mapa
        
        ⚡ **Modo rápido**: no descarga datos PML.
        """)
    elif es_solo_datos:
        st.markdown("""
        1. Pega claves de nodos
        2. Selecciona sistema, proceso, fechas
        3. (Opcional) Activa geocodificación si quieres KMZ
        4. Click en **Ejecutar**
        5. Genera y descarga lo que necesites del Centro de Descargas
        """)
    else:
        st.markdown("""
        1. Pega claves de nodos
        2. Selecciona sistema, proceso, fechas
        3. Click en **Ejecutar**
        4. Explora dashboard interactivo + BESS scoring
        
        💡 Para descargar archivos: usa modo **Solo datos**.
        """)


# ═══════════════════════════════════════════════════════════════════════
# EJECUCIÓN — guarda resultado en session_state
# ═══════════════════════════════════════════════════════════════════════
if boton and nodos:
    fecha_ini = f_ini.strftime("%Y/%m/%d")
    fecha_fin = f_fin.strftime("%Y/%m/%d")
    t0 = time.time()

    # Limpiar estado previo de descargas
    for k in ["excel_datos_bytes", "excel_analisis_bytes", "kmz_bytes"]:
        if k in st.session_state:
            st.session_state[k] = None

    try:
        acumulado = {}
        errores = []
        fx_info = {"moneda": "MXN", "tc_promedio": None, "tc_lookup": {},
                   "fuente": "N/A", "advertencia": ""}

        if necesita_datos:
            if f_ini >= f_fin:
                st.error("❌ La fecha inicial debe ser anterior a la final.")
                st.stop()

            progress = st.progress(0, text="Iniciando descarga...")
            def cb(done, total):
                progress.progress(min(done/total, 1.0) if total > 0 else 0,
                                  text=f"Descargando: {done}/{total} ({(done/total*100 if total>0 else 0):.0f}%)")

            with st.spinner("Consultando CENACE..."):
                acumulado, errores, info_sistemas, nodos_sin_sis = descargar_pml_auto(
                    nodos, fecha_ini, fecha_fin, proceso, catalogo, progress_cb=cb,
                )
            progress.progress(1.0, text="✅ Datos descargados")

            # Mostrar info de sistemas auto-detectados
            if len(info_sistemas) > 1:
                sistemas_str = " · ".join(f"{s}: {n}" for s, n in info_sistemas.items())
                st.info(f"🔄 Multi-sistema detectado: {sistemas_str}")
            if nodos_sin_sis:
                st.caption(f"⚠️ {len(nodos_sin_sis)} nodos no estaban en el catálogo. "
                           f"Se asumió SIN como fallback.")

            if not acumulado:
                st.error("❌ No se recibieron datos PML.")
                st.stop()

            # ─── CONVERSIÓN A USD si aplica ───
            moneda_seleccionada = st.session_state.get("moneda", "MXN")
            if moneda_seleccionada == "USD":
                token = obtener_token_banxico()
                if not token:
                    st.warning(
                        "⚠️ No se encontró el token de Banxico en los secrets de Streamlit. "
                        "El análisis se mostrará en MXN. Para usar USD, agrega "
                        "`BANXICO_TOKEN` en Settings → Secrets de tu app."
                    )
                    st.session_state["moneda"] = "MXN"
                    fx_info["advertencia"] = "Sin token Banxico — análisis en MXN."
                else:
                    # Descargar serie FIX para el período (con buffer de 5 días para fin de semana)
                    f_ini_buffer = (f_ini - timedelta(days=5)).strftime("%Y-%m-%d")
                    f_fin_buffer = f_fin.strftime("%Y-%m-%d")

                    with st.spinner("📥 Descargando tipo de cambio FIX de Banxico..."):
                        fx_dict = cargar_fx_banxico(f_ini_buffer, f_fin_buffer, token)

                    if not fx_dict:
                        st.warning(
                            "⚠️ No se pudo obtener el tipo de cambio de Banxico. "
                            "El análisis se mostrará en MXN."
                        )
                        st.session_state["moneda"] = "MXN"
                        fx_info["advertencia"] = "Banxico no respondió — análisis en MXN."
                    else:
                        # Construir lookup completo (incluye fines de semana)
                        fechas_unicas = set()
                        for nodo, filas in acumulado.items():
                            for f in filas:
                                fechas_unicas.add(f.get("fecha", ""))
                        fx_lookup = construir_fx_lookup(fx_dict, fechas_unicas)

                        # Aplicar conversión
                        acumulado = aplicar_conversion_usd(acumulado, fx_lookup)

                        # Guardar info
                        if fx_lookup:
                            tcs = list(fx_lookup.values())
                            fx_info = {
                                "moneda": "USD",
                                "tc_promedio": sum(tcs)/len(tcs) if tcs else None,
                                "tc_min": min(tcs) if tcs else None,
                                "tc_max": max(tcs) if tcs else None,
                                "tc_lookup": fx_lookup,
                                "fuente": "Banxico FIX (SF43718)",
                                "advertencia": "",
                            }
                            st.success(
                                f"💱 Análisis convertido a USD. TC FIX promedio del período: "
                                f"**{fx_info['tc_promedio']:.4f} MXN/USD** "
                                f"(rango: {fx_info['tc_min']:.4f} – {fx_info['tc_max']:.4f})"
                            )

        # Geocodificación si necesita mapa O si en solo datos pidió geo
        matches_df = None
        if necesita_mapa or (es_solo_datos and incluir_geo_solo_datos):
            with st.spinner("Cargando OSM (primera vez ~30s)..."):
                osm_subs = cargar_osm_subestaciones()
            if osm_subs:
                with st.spinner("Matcheando nodos con OSM..."):
                    nodos_a_matchear = list(acumulado.keys()) if necesita_datos else nodos
                    resultados = matchear_nodos(nodos_a_matchear, catalogo, osm_subs)
                    matches_df = pd.DataFrame(resultados)

        # Guardar todo en session_state
        # Sistema: si solo 1 → ese; si multi → "MULTI" + dict
        sistema_str = "MULTI" if len(info_sistemas or {}) > 1 else (
            list(info_sistemas.keys())[0] if info_sistemas else "SIN"
        )
        st.session_state["consulta_ejecutada"] = True
        st.session_state["acumulado"] = acumulado
        st.session_state["matches_df"] = matches_df
        st.session_state["fx_info"] = fx_info
        st.session_state["consulta_params"] = {
            "sistema": sistema_str, "proceso": proceso,
            "info_sistemas": info_sistemas,
            "fecha_ini": fecha_ini, "fecha_fin": fecha_fin,
            "modo": modo, "errores": errores,
            "tiempo": time.time() - t0,
            "moneda_aplicada": st.session_state.get("moneda", "MXN"),
        }

    except Exception as e:
        st.error(f"❌ Error: {type(e).__name__}: {e}")
        st.exception(e)


# ═══════════════════════════════════════════════════════════════════════
# RENDERIZAR RESULTADOS DESDE SESSION_STATE
# Esto se ejecuta cada vez que cambias algo (heatmap, BESS, etc)
# pero SIN volver a descargar datos
# ═══════════════════════════════════════════════════════════════════════
if st.session_state.get("consulta_ejecutada"):
    acumulado = st.session_state["acumulado"]
    matches_df = st.session_state["matches_df"]
    params = st.session_state["consulta_params"]
    fx_info = st.session_state.get("fx_info", {"moneda": "MXN"})
    # Modo actual del sidebar (no el guardado) — esto permite re-render instantáneo
    # al cambiar de modo sin re-ejecutar la descarga
    modo_cur = modo
    modo_consulta_original = params.get("modo", modo)

    # Detectar si parámetros que requieren re-descarga cambiaron
    moneda_actual = st.session_state.get("moneda", "MXN")
    moneda_consulta = params.get("moneda_aplicada", "MXN")

    # Si cambio el modo entre consulta y ahora pero los datos sirven
    cambio_solo_modo = (modo_cur != modo_consulta_original)

    # Cambios que SI requieren re-ejecutar
    parametros_cambiaron = []
    if necesita_datos:
        if moneda_actual != moneda_consulta:
            parametros_cambiaron.append(f"moneda ({moneda_consulta} → {moneda_actual})")
        # Comparar fechas y sistema/proceso
        f_ini_str = f_ini.strftime("%Y/%m/%d")
        f_fin_str = f_fin.strftime("%Y/%m/%d")
        if f_ini_str != params.get("fecha_ini") or f_fin_str != params.get("fecha_fin"):
            parametros_cambiaron.append("fechas")
        if proceso != params.get("proceso"):
            parametros_cambiaron.append(f"proceso ({params.get('proceso')} → {proceso})")

    # Validar si los datos actuales sirven para el modo actual
    modo_requiere_datos = ("📊" in modo_cur or "🔬" in modo_cur)
    modo_requiere_mapa = ("🗺️" in modo_cur or "🔬" in modo_cur)
    datos_disponibles = bool(acumulado)
    mapa_disponible = matches_df is not None and not matches_df.empty

    # Warning si cambiaron parámetros
    if parametros_cambiaron:
        st.warning(
            f"⚠️ **Algunos parámetros cambiaron:** {', '.join(parametros_cambiaron)}. "
            f"Los datos mostrados corresponden a la consulta anterior. "
            f"Da click en **⚡ Ejecutar** arriba para refrescar."
        )

    # Warning si modo cambió y faltan datos para ese modo
    if cambio_solo_modo:
        if modo_requiere_datos and not datos_disponibles:
            st.warning(
                f"⚠️ El modo **{modo_cur}** necesita datos PML descargados. "
                f"Da click en **⚡ Ejecutar** para descargarlos."
            )
            st.stop()
        if modo_requiere_mapa and not mapa_disponible:
            st.info(
                f"ℹ️ El modo **{modo_cur}** necesita geocodificación. "
                f"Da click en **⚡ Ejecutar** para hacer el matching geográfico."
            )

    # Mostrar moneda usada por la consulta vigente
    st.session_state["moneda"] = moneda_consulta  # forzar render con moneda de consulta

    # Banner FX si está en USD
    if moneda_consulta == "USD" and fx_info.get("tc_promedio"):
        st.info(
            f"💱 **Análisis en USD** · "
            f"TC promedio del período: **{fx_info['tc_promedio']:.4f} MXN/USD** · "
            f"Fuente: {fx_info['fuente']}"
        )

    # Solo mostrar métricas de descarga si el modo lo requiere
    if necesita_datos and acumulado:
        n_total = sum(len(f) for f in acumulado.values())
        cx, cy, cz = st.columns(3)
        cx.metric("Nodos con datos", f"{len(acumulado)}/{len(nodos) if nodos else len(acumulado)}")
        cy.metric("Total registros", f"{n_total:,}")
        cz.metric("Tiempo descarga", f"{params.get('tiempo', 0):.0f}s")

        if params.get("errores"):
            with st.expander(f"⚠️ {len(params['errores'])} errores en consultas"):
                st.dataframe(pd.DataFrame(params["errores"][:20]), use_container_width=True)

        # Aviso de Colab para consultas pesadas (multi-año + muchos nodos)
        # Umbral: > 1.3M filas (~150 nodos × 1 año equivalente)
        if n_total > 1_300_000:
            st.warning(
                f"⚠️ **Consulta pesada detectada** ({n_total:,} filas). "
                f"Para análisis multi-año con ≥150 nodos, te recomiendo usar el "
                f"**notebook Colab v64** que tiene 12 GB de RAM disponibles "
                f"vs 1 GB de Streamlit Cloud Free. La app web puede volverse lenta "
                f"o crashearse al generar gráficos pesados."
            )

    # Render según modo
    if "🗺️" in modo_cur:
        # Solo mapa
        if matches_df is not None and not matches_df.empty:
            n_mapeados = matches_df['lat'].notna().sum()
            n_consultados = len(matches_df)
            cm1, cm2, cm3 = st.columns(3)
            cm1.metric("Mapeados", f"{n_mapeados}/{n_consultados}")
            cc = matches_df['calidad'].value_counts()
            cm2.metric("🥇 Excelente / 🥈 Bueno",
                        f"{cc.get('🥇 Excelente', 0)} / {cc.get('🥈 Bueno', 0)}")
            cm3.metric("🥉 Aceptable", cc.get('🥉 Aceptable', 0))

            if n_mapeados > 0:
                fig_mapa = grafica_mapa(matches_df, color_by='ccr')
                if fig_mapa:
                    st.plotly_chart(fig_mapa, use_container_width=True,
                                    config={'scrollZoom': True, 'displayModeBar': True})
                render_leyenda_ccr(matches_df)
                render_panel_ccr(matches_df, None)

                sin_match = matches_df[matches_df['lat'].isna()]
                if not sin_match.empty:
                    with st.expander(f"❌ {len(sin_match)} nodos sin match"):
                        st.dataframe(
                            sin_match[['clave', 'nombre_cenace', 'ccr', 'estado', 'razon']],
                            use_container_width=True, hide_index=True,
                        )

    elif "📊" in modo_cur:
        # Solo datos: Centro de descargas
        if acumulado:
            render_centro_descargas(
                acumulado, catalogo,
                params["sistema"], params["proceso"],
                params["fecha_ini"], params["fecha_fin"],
                matches_df=matches_df,
                moneda=moneda_consulta,
                tc_info=fx_info,
            )

    else:
        # Modo Análisis: dashboard inline + BESS scoring (sin Excel — usar Solo Datos para descargar)
        if acumulado:
            render_dashboard(acumulado, catalogo, matches_df)


st.divider()
st.caption("⚡ CENACE PML Analyzer · Sebastian Roldan (SRF) · Recurrent Energy / Canadian Solar")
